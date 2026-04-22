import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
import numpy as np
import time
import warnings
import traceback


# Warnings ignore
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=FutureWarning)
pd.options.mode.chained_assignment = None

st.set_page_config(page_title="FRP Consolidation Tool", layout="wide")
st.markdown("""
    <style>
    /* Multiselect-ийн өндрийн хязгаарыг арилгах */
    .stMultiSelect div[data-baseweb="select"] > div:first-child {
        max-height: none !important;
        overflow-y: visible !important;
    }

    /* Сонгогдсон tag-уудын өнгийг #101D79 цэнхэр болгох */
    .stMultiSelect span[data-baseweb="tag"] {
        background-color: #101D79 !important; /* Таны хүссэн цэнхэр өнгө */
        color: white !important; /* Текстийн өнгө цагаан */
        font-size: 14px !important;
        margin: 2px !important;
        padding: 5px 10px !important;
        border-radius: 4px !important; /* Буланг үл ялиг дугуйлах */
    }

    /* Tag доторх 'X' устгах товчны өнгийг цагаан болгох */
    .stMultiSelect span[data-baseweb="tag"] svg {
        fill: white !important;
    }

    /* Multiselect-ийн ерөнхий фонтыг томруулах */
    .stMultiSelect {
        min-height: 50px;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border: none;
        color: white;
    }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        border: none;
        color: white;
    }
</style>
""", unsafe_allow_html=True)

st.title("📊 FRP Data Consolidation Tool")

# Session State Initialize
if 'final_output' not in st.session_state:
    st.session_state.final_output = None

# Sidebar
with st.sidebar:
    st.header("⚙️ Тохиргоо")
    template_file = st.file_uploader("1. Template файлаа сонгоно уу", type="xlsx")
    source_files = st.file_uploader("2. Нэгтгэх файлуудаа сонгоно уу", type="xlsx", accept_multiple_files=True)
    
    available_sheets = [
        "1.1det", "1.2det", "1.3det", "1R(i)", "1R(ii)", "1R(iii)", "1R(iv)", 
        "2R", 
        "3R", 
        "4det1", "4R", "4R(i)", 
        "5det1", "5det2", "5R(i)", "5R(ii)", 
        "6det1", "6det2", "6R", 
        "7det1", "7det2", "7det3", "7R", "7R(i)", "7R(ii)", "7R(iii)", 
        "8det1", "8det2", "8R", "8R(i)",
        "9det", "9R",
        "10det1", "10det2", "10det3",
        "Repo1", "Repo2", "Repo",
        "10R(i)", "10R(ii)", "10R(iii)", "10R(iv)", "10R(v)",
        "11det1", "11det2", "11R", "11R(i)",
        "12det", "12R",
        "13det", "13R", "13R(i)",
        "14R", "14R(i)",
        "17det1", "17det2", "17R", "17R(ii)",
        "ICT Report Raw Summary",
        "18det",
        "19R",
        "20R", "20R(i)",
        "21R", "21R(i)",
        "22R(i)", "22R(vi)",
        "24det", "24R(i)", "24R(ii)",
        "25det1", "25det2", "25R(i)", "25R(ii)",
        "26"
    ]
    # 2. "Бүгдийг сонгох" Checkbox нэмэх
# value=True гэвэл програм асах үед шууд бүгдийг нь чагталсан (сонгосон) байна.
select_all = st.checkbox("Бүх Sheet-үүдийг сонгох", value=True)

# 3. Multiselect хэсгийг нөхцөлтэйгээр өөрчлөх
if select_all:
    selected_sheets = st.multiselect(
        "3. Нэгтгэх Sheet-үүдээ сонгоно уу:", 
        available_sheets, 
        default=available_sheets
    )
else:
    selected_sheets = st.multiselect(
        "3. Нэгтгэх Sheet-үүдээ сонгоно уу:", 
        available_sheets, 
        default=[] # Checkbox-ийг болиулбал хоосон болно
    )

# Хэдэн sheet сонгогдсоныг харуулах (хяналтанд хэрэгтэй)
st.info(f"Нийт {len(selected_sheets)} sheet сонгогдсон байна.")

# Session state-д үр дүнг хадгалах
if 'is_processing' not in st.session_state:
    st.session_state.is_processing = False
if 'final_output' not in st.session_state:
    st.session_state.final_output = None

# 2. "Нэгтгэх" процесс
col1, col2 = st.columns([1, 5])
with col1:
    process_btn = st.button(
        "Нэгтгэх" if not st.session_state.is_processing else "⏳ Боловсруулж байна...",
        use_container_width=True,
        disabled=st.session_state.is_processing,
        type="primary"  # Modern primary button style
    )

if process_btn and not st.session_state.is_processing:
    if not template_file or not source_files or not selected_sheets:
        st.error("Файлууд болон Sheet-ээ бүрэн сонгоно уу!")
    else:
        with st.spinner('Боловсруулж байна...'):
            with col2:
                status_text = st.empty()
                timer_text = st.empty()
                start_time = time.time()
                error_logs = []
            
            try:
                template_file.seek(0)
                template_bytes = template_file.read()
                output_io = io.BytesIO(template_bytes)
                book = load_workbook(output_io)
                output_io.seek (0)
                template_file.seek(0)
                
                for sheet_name in selected_sheets:
                    # Дата хадгалах савнууд
                    generic_dfs = []    
                    t1_13_dfs, t2_13_dfs = [], [] 
                    t1_4d_dfs, t2_4d_dfs = [], [] 
                    t1_5d_dfs, t2_5d_dfs = [], [] 
                    t1_6d_dfs, t2_6d_dfs = [], []
                    t1_6d2_dfs, t2_6d2_dfs = [], []
                    t1_3r_dfs, t2_3r_dfs = [], []
                    t1_7d1_dfs, t2_7d1_dfs = [], []
                    t1_10d1_dfs, t2_10d1_dfs, t3_10d1_dfs, t4_10d1_dfs = [], [], [], []
                    t1_7d2_dfs = []
                    t2_7d2_dfs = []
                    t1_7d3_dfs = []
                    t1_8d1_dfs = []
                    t1_8d2_dfs, t2_8d2_dfs = [], []
                    t1_9d_dfs, t2_9d_dfs, t3_9d_dfs, t4_9d_dfs = [], [], [], []
                    t5_9d_dfs, t6_9d_dfs, t7_9d_dfs, t8_9d_dfs = [], [], [], []
                    r1_i_cols, r1_ii_cols, r1_iii_cols, r1_iv_cols = [], [], [], []
                    r2_dfs = []
                    r4_dfs = []
                    r4_i_cols = []
                    r5_i_cols, r5_ii_cols = [], []
                    r6_cols = []
                    r7_i_cols = []
                    r7_ii_data = []
                    r7iii_cols = []
                    r7_cols = []
                    r8_cols = []
                    r8i_cols = []
                    r9_cols = []
                    r1_sum_val = 0
                    det5_1_data = []
                    all_10d2_dfs = []
                    all_10d3_dfs = []
                    all_repo1_t1_dfs = []
                    all_repo1_t2_dfs = []
                    all_repo2_t1_dfs = []
                    all_repo2_t2_dfs = []
                    all_repo_cols = []
                    all_10ri_t1_dfs = []
                    all_10ri_t2_dfs = []
                    all_10rii_dfs = []
                    all_10riii_cols = []
                    all_10riv_t1_dfs = []
                    all_10riv_t2_dfs = []
                    all_10riv_t3_dfs = []
                    all_10rv_cols = []
                    all_11det1_dfs = []
                    all_11det2_dfs = []
                    all_11r_cols = []
                    all_11ri_cols = []
                    all_12det_dfs = []
                    all_12r_cols = []
                    all_13det_t1_dfs = []
                    all_13det_t2_dfs = []
                    all_13r_cols = []
                    all_13ri_cols = []
                    all_14r_cols = []
                    all_14r1_t1 = []
                    all_14r1_t2 = []
                    all_14r1_t3 = []
                    all_14r1_t4 = []
                    all_17det1_dfs = []
                    all_17det2_dfs = []
                    all_17r_cols = []
                    all_17rii_data = []
                    all_ict_summary_dfs = []
                    all_18det_cols = []
                    all_19r_cols = []
                    all_20r_cols = []
                    all_20ri_cols = []
                    all_21r_cols = []
                    all_21ri_cols = []
                    all_22ri_cols = []
                    all_22rvi_cols = []
                    all_24det_dfs = []
                    all_24ri_sum_data = []
                    all_24ri_cols = []
                    all_24rii_cols = []
                    all_25det1_dfs = []
                    all_25det2_dfs = []
                    all_25ri_cols = []
                    all_25rii_sum_data = []
                    all_25rii_cols = []
                    all_26_sum_data = []
                    all_26_cols = []

                    for file in source_files:
                        # Timer болон Status шинэчлэх
                        elapsed = time.time() - start_time
                        status_text.markdown(f"⏳ **Уншиж байна:** `{file.name}` | Sheet: `{sheet_name}`")
                        timer_text.markdown(f"🕒 **Хугацаа:** {elapsed:.1f} сек")
                        
                        file_name = file.name
                        try:
                            parts = file_name.split("_")
                            if len(parts) >= 4:
                                ac_value = parts[3].split(".")[0] # .xlsx-ийг салгах
                            else:
                                ac_value = "Unknown"
                            
                            # --- 1.1det Логик (ЗАСВАР: Багана алгасах) ---
                            if sheet_name == "1.1det":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, skiprows=7, usecols="A:Z", header=None, dtype=object, engine='openpyxl')
                                f_idx = next((i for i, v in enumerate(df_raw.iloc[:, 1]) if str(v).strip() == "ЗААВАР:"), None)
                                end = max(0, f_idx - 12) if f_idx is not None else len(df_raw)
                                df = df_raw.iloc[:end, :].copy()
                                df[1] = df[1].apply(lambda x: np.nan if str(x).strip() in ['nan', 'None', '', 'nan '] else str(x).strip())
                                df = df.dropna(subset=[1])
                                if not df.empty:
                                    df.iloc[:, 0] = str(ac_value)
                                    generic_dfs.append(df)

                            # --- 1.2det Логик ---
                            elif sheet_name == "1.2det":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, skiprows=6, usecols="A:S", header=None, dtype=object, engine='openpyxl')
                                f_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == "ЗААВАР:"), None)
                                end = max(0, f_idx - 9) if f_idx is not None else len(df_raw)
                                df = df_raw.iloc[:end, :].copy()
                                df[1] = df[1].apply(lambda x: np.nan if str(x).strip() in ['nan', 'None', '', 'nan '] else str(x).strip())
                                df = df.dropna(subset=[1])
                                if not df.empty:
                                    df.iloc[:, 0] = str(ac_value)
                                    generic_dfs.append(df)

                            # --- 1.3det Логик ---
                            elif sheet_name == "1.3det":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, skiprows=7, usecols="A:Q", header=None, dtype=object, engine='openpyxl')
                                ip_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == "B. IP"), None)
                                if ip_idx is not None:
                                    t1 = df_raw.iloc[:max(0, ip_idx-3), :].copy()
                                    t1[1] = t1[1].astype(str).str.strip().replace(['nan', 'None', '', 'nan '], np.nan)
                                    t1 = t1.dropna(subset=[1])
                                    if not t1.empty:
                                        t1.iloc[:, 0] = str(ac_value); t1.iloc[:, 14:16] = np.nan; t1_13_dfs.append(t1)
                                    t2_raw = df_raw.iloc[ip_idx+1:, :].copy()
                                    ins_idx = next((i for i, v in enumerate(t2_raw.iloc[:, 1]) if str(v).strip() == "ЗААВАР:"), None)
                                    t2 = t2_raw.iloc[:max(0, ins_idx-4), :].copy() if ins_idx is not None else t2_raw
                                    t2[1] = t2[1].apply(lambda x: np.nan if str(x).strip() in ['nan', 'None', '', 'nan '] else str(x).strip())
                                    t2 = t2.dropna(subset=[1])
                                    if not t2.empty:
                                        t2.iloc[:, 0] = str(ac_value); t2_13_dfs.append(t2)

                            # --- 1R(i) Логик ---
                            elif sheet_name == "1R(i)":
                                val_f12 = pd.read_excel(file, sheet_name=sheet_name, usecols="F", skiprows=11, nrows=1, header=None).iloc[0,0]
                                r1_sum_val += float(val_f12) if pd.notna(val_f12) else 0
                                df_r = pd.read_excel(file, sheet_name=sheet_name, usecols="F", skiprows=3, nrows=25, header=None)
                                if not df_r.empty:
                                    df_r.iloc[0, 0] = str(ac_value); r1_i_cols.append(df_r.iloc[:, 0].values)

                            # --- 1R(ii) Логик ---
                            elif sheet_name == "1R(ii)":
                                df_r = pd.read_excel(file, sheet_name=sheet_name, usecols="F", skiprows=4, nrows=26, header=None)
                                if not df_r.empty:
                                    df_r.iloc[0, 0] = str(ac_value); r1_ii_cols.append(df_r.iloc[:, 0].values)

                            # --- 1R(iii) Логик ---
                            elif sheet_name == "1R(iii)":
                                df_r = pd.read_excel(file, sheet_name=sheet_name, usecols="F", skiprows=2, nrows=10, header=None)
                                if not df_r.empty:
                                    df_r.iloc[0, 0] = str(ac_value); r1_iii_cols.append(df_r.iloc[:, 0].values)

                            # --- 1R(iv) Логик ---
                            elif sheet_name == "1R(iv)":
                                df_r = pd.read_excel(file, sheet_name=sheet_name, usecols="F", skiprows=3, nrows=24, header=None)
                                if not df_r.empty:
                                    df_r.iloc[0, 0] = str(ac_value); r1_iv_cols.append(df_r.iloc[:, 0].values)

                            # --- 2R ---
                            elif sheet_name == "2R":
                                # Сонгосон баганууд: A, B, C, D, E, F, G, H, I, J, K, O, P
                                # Индексээр (0-ээс эхэлбэл): 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 14, 15
                                c2 = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 14, 15]
                                
                                # 7-р мөрнөөс эхэлж унших (skiprows=6)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, skiprows=6, header=None, dtype=object, engine='openpyxl')
                                
                                # B баганаас (df_raw-ийн 1-р багана) "ЗААВАР:" текстийг хайх
                                ins_idx = next((i for i, v in enumerate(df_raw.iloc[:, 1]) if str(v).strip() == "ЗААВАР:"), None)
                                
                                # endRangeRow = "ЗААВАР:" мөрөөс 5-ыг хасна
                                end_row = ins_idx - 5 if ins_idx is not None else len(df_raw)
                                df = df_raw.iloc[:end_row, :].copy()
                                
                                # --- ШИНЭ ФИЛЬТЕР: B багана хоосон бол мөрийг устгах ---
                                # str(x).strip().lower() ашиглан 'nan', 'none', эсвэл хоосон зайг (' ') NaN болгоно
                                df[1] = df[1].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', '', 'nan '] else x)
                                
                                # NaN утгатай мөрүүдийг (B баганаар) бүрэн устгах
                                df = df.dropna(subset=[1])
                                
                                if not df.empty:
                                    # A баганыг (index 0) компанийн кодоор солих
                                    df.iloc[:, 0] = str(ac_value)
                                    # Зөвхөн шаардлагатай багануудыг (c2) авч хадгалах
                                    r2_dfs.append(df.iloc[:, c2])

                            # --- 3R ---
                            elif sheet_name == "3R":
                                # Сонгосон баганууд: A,B,C,D,E,F,G,H,I,J,K,L,M,O,P,R,S,U,V,Y,Z
                                # Индексээр (0-ээс эхэлбэл): 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 17, 18, 20, 21, 24, 25
                                c3 = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 17, 18, 20, 21, 24, 25]
                                
                                # 8-р мөрнөөс эхэлж унших (skiprows=7)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, skiprows=7, header=None, dtype=object, engine='openpyxl')
                                
                                # B баганаас "Joint Ventures" хайх (pandas дотор index 1)
                                split_txt = "Joint Ventures"
                                s_idx = next((i for i, v in enumerate(df_raw.iloc[:, 1]) if str(v).strip() == split_txt), None)
                                
                                # Table 1: Эхнээс s_idx хүртэл
                                if s_idx is not None:
                                    t1 = df_raw.iloc[:s_idx, :].copy()
                                    t1[1] = t1[1].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', ''] else x)
                                    t1 = t1.dropna(subset=[1])
                                    if not t1.empty:
                                        t1.iloc[:, 0] = str(ac_value)
                                        t1_3r_dfs.append(t1.iloc[:, c3])
                                        
                                    # Table 2: s_idx + 1-ээс "ЗААВАР:" хүртэл
                                    t2_raw = df_raw.iloc[s_idx+1:, :].copy()
                                    ins_idx = next((i for i, v in enumerate(t2_raw.iloc[:, 1]) if str(v).strip() == "ЗААВАР:"), None)
                                    end_row = ins_idx - 4 if ins_idx is not None else len(t2_raw)
                                    
                                    t2 = t2_raw.iloc[:end_row, :].copy()
                                    t2[1] = t2[1].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', ''] else x)
                                    t2 = t2.dropna(subset=[1])
                                    if not t2.empty:
                                        t2.iloc[:, 0] = str(ac_value)
                                        t2_3r_dfs.append(t2.iloc[:, c3])

                            # --- 4det1 Логик (ЗАСВАР: B багана устгах) ---
                            elif sheet_name == "4det1":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, skiprows=9, usecols="A:AO", header=None, dtype=object, engine='openpyxl')
                                c4 = [0, 1, 5, 6, 7, 8, 9, 10, 11, 13, 14, 15, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 31, 37, 38, 39, 40]
                                split_txt = "2. General provisions (-): 123499, 130099, 243499"
                                s_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == split_txt), None)
                                if s_idx is not None:
                                    t1 = df_raw.iloc[:s_idx, :].copy()
                                    t1[1] = t1[1].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', '', 'none '] else x)
                                    t1 = t1.dropna(subset=[1])
                                    if not t1.empty: t1.iloc[:, 0] = str(ac_value); t1_4d_dfs.append(t1.iloc[:, c4])
                                    
                                    t2_raw = df_raw.iloc[s_idx+1:, :].copy()
                                    ins_idx = next((i for i, v in enumerate(t2_raw.iloc[:, 0]) if str(v).strip() == "ЗААВАР:"), None)
                                    t2 = t2_raw.iloc[:max(0, ins_idx-18), :].copy() if ins_idx is not None else t2_raw
                                    t2[1] = t2[1].apply(lambda x: np.nan if str(x).strip() in ['nan', 'None', '', 'None '] else x)
                                    t2 = t2.dropna(subset=[1])
                                    if not t2.empty: t2.iloc[:, 0] = str(ac_value); t2_4d_dfs.append(t2.iloc[:, c4])

                            # --- 4R ---
                            elif sheet_name == "4R":
                                # Сонгосон баганууд: A, B, C, F, G, H, I, M, N
                                # Индексээр (0-ээс эхэлбэл): 0, 1, 2, 5, 6, 7, 8, 12, 13
                                c4r = [0, 1, 2, 5, 6, 7, 8, 12, 13]
                                
                                # 6-р мөрнөөс эхэлж унших (skiprows=5)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, skiprows=5, header=None, dtype=object, engine='openpyxl')
                                
                                # B баганаас (df_raw-ийн индекс 1) "ЗААВАР:" текстийг хайх
                                ins_idx = next((i for i, v in enumerate(df_raw.iloc[:, 1]) if str(v).strip() == "ЗААВАР:"), None)
                                
                                # endRangeRow = "ЗААВАР:" мөрөөс 6-г хасна
                                end_row = ins_idx - 6 if ins_idx is not None else len(df_raw)
                                df = df_raw.iloc[:end_row, :].copy()
                                
                                # --- ФИЛЬТЕР: B багана хоосон бол мөрийг устгах ---
                                df[1] = df[1].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', '', 'nan '] else x)
                                df = df.dropna(subset=[1])
                                
                                if not df.empty:
                                    # A баганыг (index 0) компанийн кодоор солих
                                    df.iloc[:, 0] = str(ac_value)
                                    # Зөвхөн шаардлагатай багануудыг (c4r) авч хадгалах
                                    r4_dfs.append(df.iloc[:, c4r])

                            # --- 4R(i) ---
                            elif sheet_name == "4R(i)":
                                # G4:G45 хүртэл унших (skiprows=3 нь 4-р мөр, nrows=42 нь 45-р мөр хүртэл)
                                df_r = pd.read_excel(file, sheet_name=sheet_name, usecols="G", skiprows=3, nrows=42, header=None)
                                
                                if not df_r.empty:
                                    # G4 нүд (index 0,0) дээрх утгыг компанийн кодоор (AC code) солих
                                    df_r.iloc[0, 0] = str(ac_value)
                                    # Уншсан багана датаг жагсаалтад нэмэх
                                    r4_i_cols.append(df_r.iloc[:, 0].values)

                            # --- 5det1 ---
                            elif sheet_name == "5det1":
                                # E3:I27 хүртэлх датаг унших (25 мөр, 5 багана)
                                # skiprows=2 нь 3-р мөрнөөс эхлэнэ, nrows=25 нь 27-р мөр хүртэл
                                df_raw = pd.read_excel(
                                    file, 
                                    sheet_name=sheet_name, 
                                    usecols="E:I",  # E, F, G, H, I багануудыг унших
                                    skiprows=2,     # 3-р мөрнөөс эхлэх
                                    nrows=25,       # 27-р мөр хүртэл (25 мөр)
                                    header=None,
                                    dtype=object,
                                    engine='openpyxl'
                                )
                                
                                if not df_raw.empty:
                                    # E3 нүд (DataFrame-ийн 0,0 байрлал) дээрх утгыг AC кодоор солих
                                    df_raw.iloc[0, 0] = str(ac_value)
                                    
                                    # Уншсан датаг хадгалах
                                    det5_1_data.append(df_raw)

                            # --- 5det2 Логик (ЗАСВАР: B багана устгах) ---
                            elif sheet_name == "5det2":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, skiprows=7, usecols="A:M", header=None, dtype=object, engine='openpyxl')
                                c5 = [0, 1, 6, 7, 8, 9, 10, 11, 12]
                                split_txt = "2. Biololgical assets - animals (161000, 251000, 251100)"
                                split_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == split_txt), None)
                                if split_idx is not None:
                                    t1 = df_raw.iloc[:max(0, split_idx - 1), :].copy()
                                    t1[1] = t1[1].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', '', 'none '] else x)
                                    t1 = t1.dropna(subset=[1])
                                    if not t1.empty: t1.iloc[:, 0] = str(ac_value); t1_5d_dfs.append(t1.iloc[:, c5])
                                    
                                    t2_raw = df_raw.iloc[split_idx + 1:, :].copy()
                                    total_idx = next((i for i, v in enumerate(t2_raw.iloc[:, 0]) if str(v).strip() == "Total"), None)
                                    t2 = t2_raw.iloc[:max(0, total_idx - 1), :].copy() if total_idx is not None else t2_raw
                                    t2[1] = t2[1].apply(lambda x: np.nan if str(x).strip() in ['nan', 'None', '', 'None '] else x)
                                    t2 = t2.dropna(subset=[1])
                                    if not t2.empty: t2.iloc[:, 0] = str(ac_value); t2_5d_dfs.append(t2.iloc[:, c5])

                            # --- 5R Sheets Логик ---
                            elif sheet_name == "5R(i)":
                                df_r = pd.read_excel(file, sheet_name=sheet_name, usecols="F", skiprows=3, nrows=28, header=None)
                                if not df_r.empty: df_r.iloc[0, 0] = str(ac_value); r5_i_cols.append(df_r.iloc[:, 0].values)
                            elif sheet_name == "5R(ii)":
                                df_r = pd.read_excel(file, sheet_name=sheet_name, usecols="F", skiprows=3, nrows=37, header=None)
                                if not df_r.empty: df_r.iloc[0, 0] = str(ac_value); r5_ii_cols.append(df_r.iloc[:, 0].values)

                            # --- 6det1 Логик ---
                            elif sheet_name == "6det1":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object, engine='openpyxl')
                                # Template-д хуулах багануудын индекс
                                c6 = [0, 1, 5, 6, 7, 8, 9, 10, 11, 14, 15, 16, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 32, 38, 39, 40, 41]
                                
                                # Хуваах цэгүүдийг олох
                                split_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == "2. General provisions (-): 123099, 243099"), None)
                                rel_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == "Relationship:"), None)
                                
                                if split_idx is not None:
                                    # Table 1: Мөр 10-аас split_idx хүртэл
                                    t1 = df_raw.iloc[9:split_idx, :].copy()
                                    # B багана (индекс 1) дээрх хоосон утга, зайг NaN болгох
                                    t1[1] = t1[1].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', '', 'none '] else x)
                                    # B багана хоосон бол мөрийг устгах
                                    t1 = t1.dropna(subset=[1])
                                    if not t1.empty:
                                        t1.iloc[:, 0] = str(ac_value)
                                        t1_6d_dfs.append(t1.iloc[:, c6])
                                        
                                    # Table 2: split_idx-ээс rel_idx хүртэл
                                    if rel_idx is not None:
                                        # rel_idx-ээс дээшх 6 мөрийг алгасаж (Нийт дүнгийн хэсэг) унших
                                        t2 = df_raw.iloc[split_idx+1 : rel_idx-6, :].copy()
                                        # B багана (индекс 1) дээрх хоосон утга, зайг NaN болгох
                                        t2[1] = t2[1].apply(lambda x: np.nan if str(x).strip() in ['nan', 'None', '', 'None '] else x)
                                        # B багана хоосон бол мөрийг устгах
                                        t2 = t2.dropna(subset=[1])
                                        if not t2.empty:
                                            t2.iloc[:, 0] = str(ac_value)
                                            t2_6d_dfs.append(t2.iloc[:, c6])
                            # --- 6det2  ---
                            elif sheet_name == "6det2":
                                # Унших баганууд: A, B, C, D, G, H, I, J, K, L
                                target_cols_6d2 = [0, 1, 2, 3, 6, 7, 8, 9, 10, 11]
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object, engine='openpyxl')
                                
                                # 1. Table 1: Салгах болон Шүүх
                                split_txt = "b. General provisions"
                                s_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == split_txt), None)
                                
                                t1 = df_raw.iloc[5:s_idx, :].copy() if s_idx is not None else pd.DataFrame()
                                # B баганыг (index 1) жижиг үсгээр шалгаж шүүх
                                t1[2] = t1[2].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', '', '"if needed, add more rows"'] else x)
                                t1 = t1.dropna(subset=[2])
                                
                                if not t1.empty:
                                    t1.iloc[:, 0] = str(ac_value) # Компанийн кодоор солих
                                    t1_6d2_dfs.append(t1.iloc[:, target_cols_6d2])

                                # 2. Table 2: Салгах болон Шүүх
                                if s_idx is not None:
                                    ins_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == "ЗААВАР:"), None)
                                    end_row = ins_idx - 4 if ins_idx is not None else len(df_raw)
                                    
                                    t2 = df_raw.iloc[s_idx+1:end_row, :].copy()
                                    # B баганыг жижиг үсгээр шалгаж шүүх (Table 2 нэмэлт шүүлтүүртэй)
                                    t2[2] = t2[2].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', '', 'general provisions', '"if needed, add more rows"'] else x)
                                    t2 = t2.dropna(subset=[2])
                                    
                                    if not t2.empty:
                                        t2.iloc[:, 0] = str(ac_value)
                                        t2_6d2_dfs.append(t2.iloc[:, target_cols_6d2])
                            # --- 6R ---
                            elif sheet_name == "6R":
                                # data_only=True ашиглан томьёоны утгыг (value) унших
                                df_r = pd.read_excel(file, sheet_name=sheet_name, usecols="E", skiprows=3, nrows=53, header=None)
                                
                                if not df_r.empty:
                                    # E4 нүд (DataFrame-ийн 0,0 байрлал) дээрх утгыг AC кодоор солих
                                    df_r.iloc[0, 0] = str(ac_value)
                                    # Уншсан багана датаг жагсаалтад нэмэх
                                    r6_cols.append(df_r.iloc[:, 0].values)

                            # --- 7det1 ---
                            elif sheet_name == "7det1":
                                # Унших баганууд: A, B, F, G, H, I, J, K, M, N, V, W, X, Y + AA, AB + AZ:BF
                                # Шинэ индексүүд: 
                                # [0, 1, 5, 6, 7, 8, 9, 10, 12, 13, 21, 22, 23, 24, 26, 27, 51, 52, 53, 54, 55, 56, 57]
                                c71 = [0, 1, 5, 6, 7, 8, 9, 10, 12, 13, 21, 22, 23, 24, 26, 27, 51, 52, 53, 54, 55, 56, 57]
                                
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object, engine='openpyxl')
                                
                                # Split текст хайх
                                split_txt = "2. General provisions (-): 120199, 121099, 122099,123199, 240199, 242099, 243199"
                                s_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == split_txt), None)
                                
                                # --- Table 1 Logic ---
                                t1 = df_raw.iloc[9:s_idx, :].copy() if s_idx is not None else pd.DataFrame()
                                if not t1.empty:
                                    # B багана (index 1) хоосон бол устгах
                                    t1[1] = t1[1].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', ''] else x)
                                    t1 = t1.dropna(subset=[1])
                                    
                                    if not t1.empty:
                                        t1.iloc[:, 0] = str(ac_value) # Компани кодоор солих
                                        # Сонгосон багануудыг шүүж аваад жагсаалтад нэмэх
                                        t1_7d1_dfs.append(t1.iloc[:, c71])

                                # --- Table 2 Logic ---
                                if s_idx is not None:
                                    rel_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == "Relationship:"), None)
                                    end_row = rel_idx - 4 if rel_idx is not None else len(df_raw)
                                    
                                    t2 = df_raw.iloc[s_idx+1:end_row, :].copy()
                                    if not t2.empty:
                                        # B багана хоосон бол устгах
                                        t2[1] = t2[1].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', ''] else x)
                                        t2 = t2.dropna(subset=[1])
                                        
                                        if not t2.empty:
                                            t2.iloc[:, 0] = str(ac_value) # Компани кодоор солих
                                            # Сонгосон багануудыг шүүж аваад жагсаалтад нэмэх
                                            t2_7d1_dfs.append(t2.iloc[:, c71])

                            # --- 7det2 ---
                            elif sheet_name == "7det2":
                                # Унших баганууд: A, B, C, D, H, I, J, K, L, M (Нийт 10 багана)
                                # Index: 0, 1, 2, 3, 7, 8, 9, 10, 11, 12
                                target_cols_7d2 = [0, 1, 2, 3, 7, 8, 9, 10, 11, 12]
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object, engine='openpyxl')
                                
                                # Split текст хайх
                                split_txt = "b. General provisions"
                                s_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == split_txt), None)
                                
                                # --- Table 1 Logic ---
                                # Range: A6-аас (index 5) split_txt хүртэл
                                t1 = df_raw.iloc[5:s_idx, :].copy() if s_idx is not None else pd.DataFrame()
                                
                                # D багана (index 3) хоосон бол устгах
                                t1[3] = t1[3].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', ''] else x)
                                t1 = t1.dropna(subset=[3])
                                
                                if not t1.empty:
                                    t1.iloc[:, 0] = str(ac_value) # A column-ийг компанийн кодоор солих
                                    t1_7d2_dfs.append(t1.iloc[:, target_cols_7d2])

                                # --- Table 2 Logic ---
                                if s_idx is not None:
                                    # "ЗААВАР:" текст хайх
                                    ins_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == "ЗААВАР:"), None)
                                    # endRow = ЗААВАР-аас 4 мөр дээш
                                    end_row = ins_idx - 4 if ins_idx is not None else len(df_raw)
                                    
                                    # Range: split_txt-ээс 1 мөр доороос (s_idx+1) end_row хүртэл
                                    t2 = df_raw.iloc[s_idx+1:end_row, :].copy()
                                    
                                    # D багана (index 3) хоосон бол устгах
                                    t2[3] = t2[3].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', ''] else x)
                                    t2 = t2.dropna(subset=[3])
                                    
                                    if not t2.empty:
                                        t2.iloc[:, 0] = str(ac_value) # A column-ийг компанийн кодоор солих
                                        t2_7d2_dfs.append(t2.iloc[:, target_cols_7d2])

                            # --- 7det3 ---
                            elif sheet_name == "7det3":
                                # Унших багануудын индекс (A, B, C, D, E, F, G, H, J, K, M, N, P, Q, S, T, V, W, Y, Z)
                                # Index: 0, 1, 2, 3, 4, 5, 6, 7, 9, 10, 12, 13, 15, 16, 18, 19, 21, 22, 24, 25
                                c73 = [0, 1, 2, 3, 4, 5, 6, 7, 9, 10, 12, 13, 15, 16, 18, 19, 21, 22, 24, 25]
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object, engine='openpyxl')
                                
                                # "ЗААВАР:" текст хайх
                                ins_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == "ЗААВАР:"), None)
                                end_row = ins_idx - 5 if ins_idx is not None else len(df_raw)
                                
                                # Range: A9 (index 8) -аас end_row хүртэл
                                t1 = df_raw.iloc[8:end_row, :].copy()
                                
                                # C багана (index 2) хоосон эсвэл "Шаардлагатай бол нэмнэ үү..." байвал устгах
                                filter_vals = ['nan', 'none', '', 'шаардлагатай бол нэмнэ үү...']
                                t1[2] = t1[2].apply(lambda x: np.nan if str(x).strip().lower() in filter_vals else x)
                                t1 = t1.dropna(subset=[2])
                                
                                if not t1.empty:
                                    t1.iloc[:, 0] = str(ac_value) # Компани кодоор солих
                                    t1_7d3_dfs.append(t1.iloc[:, c73])

                            # --- 7R ---
                            elif sheet_name == "7R":
                                # F баганыг унших (Excel-ийн F3:F62 -> usecols="F", skiprows=2, nrows=60)
                                # data_only=True-тэй адил утгыг нь авахын тулд
                                df_r7 = pd.read_excel(file, sheet_name=sheet_name, usecols="F", skiprows=2, nrows=60, header=None)
                                
                                if not df_r7.empty:
                                    # F3 нүд (DataFrame-ийн 0,0 байрлал) дээрх утгыг AC кодоор солих
                                    df_r7.iloc[0, 0] = str(ac_value)
                                    # Уншсан багана датаг жагсаалтад нэмэх
                                    r7_cols.append(df_r7.iloc[:, 0].values)

                            # --- 7R(i) ---
                            elif sheet_name == "7R(i)":
                                # E2:E41 хүртэл унших (skiprows=1 нь 2-р мөр, nrows=40 нь 41-р мөр хүртэл)
                                df_r = pd.read_excel(file, sheet_name=sheet_name, usecols="E", skiprows=1, nrows=40, header=None)
                                
                                if not df_r.empty:
                                    # E2 нүд (index 0,0) дээрх утгыг AC кодоор солих
                                    df_r.iloc[0, 0] = str(ac_value)
                                    # Уншсан багана датаг жагсаалтад нэмэх
                                    r7_i_cols.append(df_r.iloc[:, 0].values)

                            # --- 7R(ii) ---
                            elif sheet_name == "7R(ii)":
                                # 5-15 мөрийг унших (skiprows=4 нь 5-р мөр, nrows=11 нь 15-р мөр хүртэл)
                                # E, I, M багануудыг унших (индекс: 4, 8, 12)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, usecols="E,I,M", skiprows=4, nrows=11, header=None)
                                
                                if not df_raw.empty:
                                    # 6-р мөр дэх утгуудыг (DataFrame-ийн индекс 1) AC кодоор солих
                                    df_raw.iloc[1, 0] = str(ac_value) # E6
                                    df_raw.iloc[1, 1] = str(ac_value) # I6
                                    df_raw.iloc[1, 2] = str(ac_value) # M6
                                    
                                    # Тухайн компанийн 3 баганыг багц болгож нэмэх
                                    r7_ii_data.append({
                                        'col_E': df_raw.iloc[:, 0].values,
                                        'col_I': df_raw.iloc[:, 1].values,
                                        'col_M': df_raw.iloc[:, 2].values
                                    })

                            # --- 7R(iii) ---
                            elif sheet_name == "7R(iii)":
                                # D баганыг унших (Excel-ийн D4:D35 -> usecols="D", skiprows=3, nrows=32)
                                df_r7iii = pd.read_excel(file, sheet_name=sheet_name, usecols="D", skiprows=3, nrows=32, header=None)
                                
                                if not df_r7iii.empty:
                                    # D4 нүд (DataFrame-ийн 0,0 байрлал) дээрх утгыг AC кодоор солих
                                    df_r7iii.iloc[0, 0] = str(ac_value)
                                    # Уншсан багана датаг жагсаалтад нэмэх
                                    r7iii_cols.append(df_r7iii.iloc[:, 0].values)

                            # --- 8det1 ---
                            elif sheet_name == "8det1":
                                # Унших баганууд: A, B, G, H, I, J, K, L
                                # Индекс: 0, 1, 6, 7, 8, 9, 10, 11 (9 нэмэгдсэн)
                                c81 = [0, 1, 6, 7, 8, 9, 10, 11]
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object, engine='openpyxl')
                                
                                # "ЗААВАР:" текст хайх
                                ins_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == "ЗААВАР:"), None)
                                # endRow = ЗААВАР-аас 27 мөр хасах
                                end_row = ins_idx - 27 if ins_idx is not None else len(df_raw)
                                
                                # Range: A7 (index 6) -аас end_row хүртэл
                                t1 = df_raw.iloc[6:end_row, :].copy()
                                
                                # B багана (index 1) хоосон бол устгах
                                t1[1] = t1[1].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', ''] else x)
                                t1 = t1.dropna(subset=[1])
                                
                                if not t1.empty:
                                    t1.iloc[:, 0] = str(ac_value) # A баганыг компанийн кодоор солих
                                    # Сонгосон c81 багануудыг хадгалах
                                    t1_8d1_dfs.append(t1.iloc[:, c81])

                            # --- 8det2 ---
                            elif sheet_name == "8det2":
                                # Унших баганууд: A:S (Индекс 0-оос 18 хүртэл)
                                target_cols_8d2 = list(range(19))
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object, engine='openpyxl')
                                
                                # Split текст болон ЗААВАР текст хайх
                                split_txt = "Төслийн дугааргүй, мөн тохируулгын данстай төслүүд"
                                s_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == split_txt), None)
                                ins_idx = next((i for i, v in enumerate(df_raw.iloc[:, 0]) if str(v).strip() == "ЗААВАР:"), None)
                                
                                # --- Table 1 Logic ---
                                # Range: A7 (index 6) -аас split_txt хүртэл
                                t1 = df_raw.iloc[6:s_idx, :].copy() if s_idx is not None else pd.DataFrame()
                                # C багана (index 2) хоосон бол устгах
                                t1[2] = t1[2].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', ''] else x)
                                t1 = t1.dropna(subset=[2])
                                
                                if not t1.empty:
                                    t1.iloc[:, 0] = str(ac_value) # Компани кодоор солих
                                    t1_8d2_dfs.append(t1.iloc[:, target_cols_8d2])

                                # --- Table 2 Logic ---
                                if s_idx is not None:
                                    end_row_t2 = ins_idx - 4 if ins_idx is not None else len(df_raw)
                                    # Range: split_txt-ээс (s_idx+1) ЗААВАР-4 хүртэл
                                    t2 = df_raw.iloc[s_idx+1:end_row_t2, :].copy()
                                    
                                    # C багана хоосон мөн "Шаардлагатай бол нэмнэ үү" байвал устгах
                                    filter_vals = ['nan', 'none', '', 'шаардлагатай бол нэмнэ үү']
                                    t2[2] = t2[2].apply(lambda x: np.nan if str(x).strip().lower() in filter_vals else x)
                                    t2 = t2.dropna(subset=[2])
                                    
                                    if not t2.empty:
                                        t2.iloc[:, 0] = str(ac_value)
                                        t2_8d2_dfs.append(t2.iloc[:, target_cols_8d2])

                            # --- 8R ---
                            elif sheet_name == "8R":
                                # E баганыг унших (Excel-ийн E2:E21 -> usecols="E", skiprows=1, nrows=20)
                                df_r8 = pd.read_excel(file, sheet_name=sheet_name, usecols="E", skiprows=1, nrows=20, header=None)
                                
                                if not df_r8.empty:
                                    # E2 нүд (DataFrame-ийн 0,0 байрлал) дээрх утгыг AC кодоор солих
                                    df_r8.iloc[0, 0] = str(ac_value)
                                    # Уншсан багана датаг жагсаалтад нэмэх
                                    r8_cols.append(df_r8.iloc[:, 0].values)

                            # --- 8R(i) ---
                            elif sheet_name == "8R(i)":
                                # D баганыг унших (Excel-ийн D5:D40 -> usecols="D", skiprows=4, nrows=36)
                                df_r8i = pd.read_excel(file, sheet_name=sheet_name, usecols="D", skiprows=4, nrows=36, header=None)
                                
                                if not df_r8i.empty:
                                    # D5 нүд (DataFrame-ийн 0,0 байрлал) дээрх утгыг AC кодоор солих
                                    df_r8i.iloc[0, 0] = str(ac_value)
                                    # Уншсан багана датаг жагсаалтад нэмэх
                                    r8i_cols.append(df_r8i.iloc[:, 0].values)

                            # --- 9det ---
                            elif sheet_name == "9det":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object, engine='openpyxl')
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                b_col = df_raw.iloc[:, 1].astype(str).str.strip()
                                
                                # Текстээр индексүүдийг олох
                                idx_b = next((i for i, v in enumerate(a_col) if v == "B. Bank accounts - Харилцах данс (110100,110109,110110)"), None)
                                idx_c = next((i for i, v in enumerate(a_col) if v == "C. Cash equivalents: Мөнгөтэй адилтгах бусад хөрөнгө"), None)
                                idx_119 = next((i for i, v in enumerate(a_col) if v == "119000, 119100 (Хязгаарлагдсан мөнгө)"), None)
                                idx_110199 = next((i for i, v in enumerate(a_col) if v == "110199 (Үнэгүйдэл)"), None)
                                idx_total = next((i for i, v in enumerate(a_col) if v == "Total cash and cash equivalents"), None)
                                idx_111200 = next((i for i, v in enumerate(a_col) if v == "111200 (4-12 сарын хугацаатай)"), None)
                                idx_112400 = next((i for i, v in enumerate(a_col) if v == "112400 (12-с дээш сарын хугацаатай)"), None)
                                idx_1192 = next((i for i, v in enumerate(a_col) if v == "119200, 119400 (Хязгаарлагдсан мөнгө)"), None)
                                idx_instr = next((i for i, v in enumerate(b_col) if v == "ЗААВАР:"), None)

                                filter_vals = ['nan', 'none', '', 'шаардлагатай бол нэмнэ үү...']

                                def process_9det_table(df_part, cols_count):
                                    if df_part.empty: return None
                                    
                                    # B багана (Index 1) хоосон утгуудыг цэвэрлэх
                                    df_part[1] = df_part[1].apply(lambda x: np.nan if str(x).strip().lower() in filter_vals else x)
                                    df_part = df_part.dropna(subset=[1])
                                    
                                    # ШҮҮЛТҮҮР: B багана дээр "Account number /бүтнээр/" байвал устгах
                                    if 1 in df_part.columns:
                                        df_part = df_part[df_part[1].astype(str).str.strip() != "Account number /бүтнээр/"]

                                    if not df_part.empty:
                                        # АЛДАА ЗАСАХ: Баганын тоог cols_count-д хүргэж бэлдэх (iloc enlarge алдаанаас сэргийлнэ)
                                        if df_part.shape[1] < cols_count:
                                            df_part = df_part.reindex(columns=range(cols_count))
                                        
                                        # Одоо A багана руу Компани кодыг аюулгүй бичнэ
                                        df_part.iloc[:, 0] = str(ac_value)
                                        return df_part.iloc[:, :cols_count]
                                    return None

                                # Table Ranges & Processing
                                if idx_b is not None:
                                    t1_9d_dfs.append(process_9det_table(df_raw.iloc[5:idx_b-3+1], 8)) # A:H (8 cols)
                                    
                                    if idx_c is not None:
                                        t2_9d_dfs.append(process_9det_table(df_raw.iloc[idx_b+3:idx_c-3+1], 15)) # A:O (15 cols)
                                        
                                        if idx_119 is not None:
                                            t3_9d_dfs.append(process_9det_table(df_raw.iloc[idx_c-3+5:idx_119-1+1], 20)) # A:T
                                            t4_9d_dfs.append(process_9det_table(df_raw.iloc[idx_119:idx_110199-1+1], 20)) # A:T
                                            
                                            if idx_110199 is not None:
                                                t5_9d_dfs.append(process_9det_table(df_raw.iloc[idx_110199-1+2:idx_total-4+1], 20))
                                                
                                if idx_111200 is not None and idx_112400 is not None:
                                    t6_9d_dfs.append(process_9det_table(df_raw.iloc[idx_111200+1:idx_112400-1+1], 20))
                                    t7_9d_dfs.append(process_9det_table(df_raw.iloc[idx_112400-1+2:idx_1192-1+1], 20))
                                    if idx_instr is not None:
                                        t8_9d_dfs.append(process_9det_table(df_raw.iloc[idx_1192-1+2:idx_instr-4+1], 20))

                                # Clean None values from all lists (t1-ээс t8 хүртэл)
                                t1_9d_dfs = [x for x in t1_9d_dfs if x is not None]
                                t2_9d_dfs = [x for x in t2_9d_dfs if x is not None]
                                t3_9d_dfs = [x for x in t3_9d_dfs if x is not None]
                                t4_9d_dfs = [x for x in t4_9d_dfs if x is not None]
                                t5_9d_dfs = [x for x in t5_9d_dfs if x is not None]
                                t6_9d_dfs = [x for x in t6_9d_dfs if x is not None]
                                t7_9d_dfs = [x for x in t7_9d_dfs if x is not None]
                                t8_9d_dfs = [x for x in t8_9d_dfs if x is not None]

                            # --- 9R ---
                            elif sheet_name == "9R":
                                # F4:F42 хүртэл унших (skiprows=3 нь 4-р мөр, nrows=39 нь 42-р мөр хүртэл)
                                df_r = pd.read_excel(file, sheet_name=sheet_name, usecols="F", skiprows=3, nrows=39, header=None)
                                
                                if not df_r.empty:
                                    # F4 нүд (index 0,0) дээрх утгыг AC кодоор солих
                                    df_r.iloc[0, 0] = str(ac_value)
                                    # Уншсан багана датаг жагсаалтад нэмэх
                                    r9_cols.append(df_r.iloc[:, 0].values)

                            # --- 10det1 ---
                            elif sheet_name == "10det1":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                
                                # Багануудын индексүүдийг тодорхойлох
                                # A, B, F:L, O:R, U:AF, AK:AY, BB:BG
                                cols_idx = [0, 1] + list(range(5, 12)) + list(range(14, 18)) + \
                                           list(range(20, 32)) + list(range(36, 51)) + list(range(53, 59))

                                # Текстээр заагч Row-уудыг олох
                                idx_non_curr = next((i for i, v in enumerate(a_col) if "Non-current (360201-360297" in v), None)
                                idx_fair_val = next((i for i, v in enumerate(a_col) if "B. Other financial liabilities measured at fair value" in v), None)
                                idx_non_curr2 = next((i for i, v in enumerate(a_col) if "Non-current (365000-365098)" in v), None)
                                idx_curr_borr = next((i for i, v in enumerate(a_col) if "Current borrowings" in v), None)

                                def process_10d1_table(df_part):
                                    if df_part.empty: return None
                                    # B багана (index 1) хоосон бол устгах
                                    df_part[1] = df_part[1].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', ''] else x)
                                    df_part = df_part.dropna(subset=[1])
                                    if not df_part.empty:
                                        df_part.iloc[:, 0] = str(ac_value) # A баганыг Компани кодоор солих
                                        return df_part.iloc[:, cols_idx]
                                    return None

                                # Table 1: A13 (index 12) -> Non-current - 1
                                if idx_non_curr:
                                    res = process_10d1_table(df_raw.iloc[12:idx_non_curr])
                                    if res is not None: t1_10d1_dfs.append(res)

                                # Table 2: Non-current -> Fair Value - 1
                                if idx_non_curr and idx_fair_val:
                                    res = process_10d1_table(df_raw.iloc[idx_non_curr:idx_fair_val])
                                    if res is not None: t2_10d1_dfs.append(res)

                                # Table 3: Fair Value + 2 -> Non-current(365...) - 1
                                if idx_fair_val and idx_non_curr2:
                                    res = process_10d1_table(df_raw.iloc[idx_fair_val+2:idx_non_curr2])
                                    if res is not None: t3_10d1_dfs.append(res)

                                # Table 4: Non-current(365...) + 1 -> Current borrowings - 1
                                if idx_non_curr2 and idx_curr_borr:
                                    res = process_10d1_table(df_raw.iloc[idx_non_curr2+1:idx_curr_borr])
                                    if res is not None: t4_10d1_dfs.append(res)

                            #--- 10det2 ---
                            elif sheet_name == "10det2":
                                # A:M баганыг унших (Индекс 0-ээс 12 хүртэл)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # C баганаас (Индекс 2) "ЗААВАР:" текстийг хайх
                                c_col = df_raw.iloc[:, 2].astype(str).str.strip()
                                end_idx = next((i for i, v in enumerate(c_col) if "ЗААВАР:" in v), None)
                                
                                if end_idx is not None:
                                    # DataReadEndRow = Row - 4 (Python индексээр 6-аас эхлэн end_idx-3 хүртэл)
                                    df_part = df_raw.iloc[6:end_idx-3, 0:13].copy() 
                                    
                                    # ШҮҮЛТҮҮР: B (index 1) болон C (index 2) багана хоосон биш байх
                                    df_part = df_part.dropna(subset=[1, 2])
                                    df_part = df_part[
                                        (df_part.iloc[:, 1].astype(str).str.strip() != "") & 
                                        (df_part.iloc[:, 2].astype(str).str.strip() != "")
                                    ]
                                    
                                    if not df_part.empty:
                                        # A баганыг (index 0) Компани кодоор солих
                                        df_part.iloc[:, 0] = str(ac_value)
                                        all_10d2_dfs.append(df_part)

                            # --- 10det3 ---
                            elif sheet_name == "10det3":
                                # A:L баганыг унших (Индекс 0-ээс 11 хүртэл)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # A баганаас (Индекс 0) "ЗААВАР:" текстийг хайх
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                end_idx = next((i for i, v in enumerate(a_col) if "ЗААВАР:" in v), None)
                                
                                if end_idx is not None:
                                    # DataReadEndRow = Row - 4 (Python индексээр 6-аас эхлэн end_idx-3 хүртэл)
                                    df_part = df_raw.iloc[6:end_idx-3, 0:12].copy() 
                                    
                                    # НӨХЦӨЛ: C (index 2) багана хоосон биш байх
                                    df_part[2] = df_part[2].apply(lambda x: np.nan if str(x).strip().lower() in ['nan', 'none', ''] else x)
                                    df_part = df_part.dropna(subset=[2])
                                    
                                    if not df_part.empty:
                                        # A баганыг (index 0) Компани кодоор replace хийх
                                        df_part.iloc[:, 0] = str(ac_value)
                                        all_10d3_dfs.append(df_part)

                            # --- Repo1 ---
                            elif sheet_name == "Repo1":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                b_col = df_raw.iloc[:, 1].astype(str).str.strip()
                                
                                # --- Table 1 Logic ---
                                # Баганууд: A:B(0:1), F:J(5:9), M:N(12:13), S:Z(18:25), AB(27), AH:AK(33:36)
                                t1_cols = [0, 1] + list(range(5, 10)) + [12, 13] + list(range(18, 26)) + [27] + list(range(33, 37))
                                
                                t1_end_idx = next((i for i, v in enumerate(a_col) if "2. General provisions (-): 123699, 243699" in v), None)
                                
                                if t1_end_idx is not None:
                                    # A10 (index 9) -аас t1_end_idx хүртэл
                                    df_t1 = df_raw.iloc[9:t1_end_idx, :].copy()
                                    # Нөхцөл: B багана (index 1) хоосон биш байх
                                    df_t1 = df_t1[df_t1.iloc[:, 1].notna() & (df_t1.iloc[:, 1].astype(str).str.strip() != "")]
                                    
                                    if not df_t1.empty:
                                        df_t1.iloc[:, 0] = str(ac_value) # A-г Компани кодоор солих
                                        all_repo1_t1_dfs.append(df_t1.iloc[:, t1_cols])

                                # --- Table 2 Logic ---
                                # Баганууд: A:B(0:1), F:J(5:9), M:N(12:13), AB(27), AH:AI(33:34)
                                t2_cols = [0, 1] + list(range(5, 10)) + [12, 13] + [27] + list(range(33, 35))
                                
                                # Start: Table1EndRow + 1 (Python индексээр бол шууд t1_end_idx + 1 орчим)
                                t2_start_row = t1_end_idx + 1 if t1_end_idx is not None else 97 # Default backup
                                t2_end_search_idx = next((i for i, v in enumerate(b_col) if "Currency" in v), None)
                                
                                if t2_end_search_idx is not None:
                                    t2_end_row = t2_end_search_idx - 3 # Row-4 гэдэг нь индексээр -3 орчим
                                    df_t2 = df_raw.iloc[t2_start_row:t2_end_row, :].copy()
                                    
                                    # Нөхцөл: B багана (index 1) хоосон биш байх
                                    df_t2 = df_t2[df_t2.iloc[:, 1].notna() & (df_t2.iloc[:, 1].astype(str).str.strip() != "")]
                                    
                                    if not df_t2.empty:
                                        df_t2.iloc[:, 0] = str(ac_value)
                                        all_repo1_t2_dfs.append(df_t2.iloc[:, t2_cols])

                            # --- Repo 2 ---
                            elif sheet_name == "Repo2":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                
                                # Багануудын индекс: A:B(0:1), F:J(5:9), M:P(12:15), S:Z(18:25), AE:AJ(30:35), AL:AQ(37:42)
                                common_cols = [0, 1] + list(range(5, 10)) + list(range(12, 16)) + \
                                              list(range(18, 26)) + list(range(30, 36)) + list(range(37, 43))
                                
                                # --- Table 1 Logic ---
                                t1_end_idx = next((i for i, v in enumerate(a_col) if "Non-current (361100, 361197, 362089)" in v), None)
                                
                                if t1_end_idx is not None:
                                    # A13 (index 12) -аас t1_end_idx хүртэл
                                    df_t1 = df_raw.iloc[12:t1_end_idx, :].copy()
                                    # Нөхцөл: B багана (index 1) хоосон биш
                                    df_t1 = df_t1[df_t1.iloc[:, 1].notna() & (df_t1.iloc[:, 1].astype(str).str.strip() != "")]
                                    
                                    if not df_t1.empty:
                                        df_t1.iloc[:, 0] = str(ac_value) # A-г Компани кодоор солих
                                        all_repo2_t1_dfs.append(df_t1.iloc[:, common_cols])

                                # --- Table 2 Logic ---
                                # Start: Table1EndRow + 2 (Python индексээр t1_end_idx + 1)
                                t2_start_row = t1_end_idx + 1 if t1_end_idx is not None else 35
                                t2_end_search_idx = next((i for i, v in enumerate(a_col) if "Total" in v), None)
                                
                                if t2_end_search_idx is not None:
                                    t2_end_row = t2_end_search_idx - 2 # Row-3 гэдэг нь индексээр -2
                                    df_t2 = df_raw.iloc[t2_start_row:t2_end_row, :].copy()
                                    
                                    # Нөхцөл: B багана (index 1) хоосон биш
                                    df_t2 = df_t2[df_t2.iloc[:, 1].notna() & (df_t2.iloc[:, 1].astype(str).str.strip() != "")]
                                    
                                    if not df_t2.empty:
                                        df_t2.iloc[:, 0] = str(ac_value)
                                        all_repo2_t2_dfs.append(df_t2.iloc[:, common_cols])

                            # --- Repo ---
                            elif sheet_name == "Repo":
                                # E4:E38 хүрээг унших (Python индексээр Row 3:38, Col 4)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # E4:E38 датаг авах (3-р мөрнөөс 38-р мөр хүртэл, 4-р багана)
                                # pandas.iloc-д төгсгөлийн индекс орохгүй тул 3:38 гэж авна
                                repo_col_data = df_raw.iloc[3:38, 4].tolist()
                                
                                if repo_col_data:
                                    # E4 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    repo_col_data[0] = str(ac_value)
                                    all_repo_cols.append(repo_col_data)

                            # --- 10R(i) ---
                            elif sheet_name == "10R(i)":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                b_col = df_raw.iloc[:, 1].astype(str).str.strip()
                                
                                # --- Table 1 Logic ---
                                t1_end_idx = next((i for i, v in enumerate(b_col) if "Current borrowing amount" in v), None)
                                
                                if t1_end_idx is not None:
                                    # A7 (index 6) -оос t1_end_idx хүртэл, A:I (0:9) багана
                                    df_t1 = df_raw.iloc[6:t1_end_idx, 0:9].copy()
                                    
                                    # НӨХЦӨЛ: B хоосон биш мөн "0"-ээс өөр байх
                                    df_t1 = df_t1[
                                        df_t1.iloc[:, 1].notna() & 
                                        (df_t1.iloc[:, 1].astype(str).str.strip() != "") & 
                                        (df_t1.iloc[:, 1].astype(str).str.strip() != "0")
                                    ]
                                    
                                    if not df_t1.empty:
                                        df_t1.iloc[:, 0] = str(ac_value) # A-г Компани кодоор солих
                                        all_10ri_t1_dfs.append(df_t1)

                                # --- Table 2 Logic ---
                                # Зөвхөн B баганаас (b_col) текстийг хайж индекс тогтоох
                                t2_end_idx = next((i for i, v in enumerate(b_col) if "Non-current borrowing amount" in v), None)
                                
                                if t2_end_idx is not None:
                                    # Table 2-ийн эхлэх мөр: t1_end_idx + 3 (Excel-ийн бүтцээс хамаарч)
                                    # Хэрэв t1_end_idx байхгүй бол статикаар 23-р мөрнөөс хайж эхэлнэ
                                    actual_start = t1_end_idx + 3 if t1_end_idx is not None else 23
                                    
                                    # Датаг унших (A:I буюу 0:9 багана)
                                    df_t2 = df_raw.iloc[actual_start:t2_end_idx, 0:9].copy()
                                    
                                    if not df_t2.empty:
                                        # НӨХЦӨЛ: B багана (Index 1) хоосон биш, мөн "0" биш байх
                                        df_t2 = df_t2[
                                            df_t2.iloc[:, 1].notna() & 
                                            (df_t2.iloc[:, 1].astype(str).str.strip() != "") & 
                                            (df_t2.iloc[:, 1].astype(str).str.strip() != "0")
                                        ]
                                        
                                        if not df_t2.empty:
                                            # 'iloc cannot enlarge' алдаанаас сэргийлж баганын тоог нөхөх
                                            if df_t2.shape[1] < 9:
                                                df_t2 = df_t2.reindex(columns=range(9))
                                            
                                            # A багана (Index 0) руу Компани кодыг бичих
                                            df_part_final = df_t2.copy()
                                            df_part_final.iloc[:, 0] = str(ac_value)
                                            all_10ri_t2_dfs.append(df_part_final)

                            # --- 10R(ii) ---
                            elif sheet_name == "10R(ii)":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # A баганаас "ЗААВАР:" хайх
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                end_idx = next((i for i, v in enumerate(a_col) if "ЗААВАР:" in v), None)
                                
                                if end_idx is not None:
                                    # Дата унших төгсгөл (ЗААВАР-аас дээшх мөрүүд)
                                    # Хэрэв зай бага бол шууд end_idx хүртэл авна
                                    read_end_row = end_idx - 5 if (end_idx - 5) > 9 else end_idx
                                    
                                    # Датаг унших (Баганыг хязгаарлахгүйгээр бүгдийг авна)
                                    df_part = df_raw.iloc[9:read_end_row, :].copy()
                                    
                                    if not df_part.empty:
                                        # ШҮҮЛТҮҮР: A багана (Index 0) хоосон биш байх
                                        df_part = df_part[df_part.iloc[:, 0].notna()]
                                        df_part = df_part[df_part.iloc[:, 0].astype(str).str.strip() != ""]
                                        
                                        if not df_part.empty:
                                            # ЧУХАЛ: J багана (Index 9) байхгүй бол нэмж үүсгэнэ
                                            # Ингэснээр J багана хоосон байсан ч Pandas-т 10 баганатай болгож харуулна
                                            if df_part.shape[1] < 10:
                                                df_part = df_part.reindex(columns=range(10))
                                            
                                            # Одоо J багана (Index 9) руу компанийн кодыг бичнэ
                                            df_part.iloc[:, 9] = str(ac_value)
                                            
                                            # Зөвхөн A:J (0:10) багануудыг нэгтгэх жагсаалт руу нэмнэ
                                            all_10rii_dfs.append(df_part.iloc[:, :10])

                            # --- 10R(iii) ---
                            elif sheet_name == "10R(iii)":
                                # D4:D29 хүрээг унших (Python индексээр Row 3:29, Col 3)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # D4:D29 датаг авах (3-р мөрнөөс 29-р мөр хүртэл, 3-р багана)
                                # pandas.iloc[3:29, 3] гэдэг нь Excel-ийн D4:D29 юм
                                r3_col_data = df_raw.iloc[3:29, 3].tolist()
                                
                                if r3_col_data:
                                    # D4 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r3_col_data[0] = str(ac_value)
                                    all_10riii_cols.append(r3_col_data)

                            # --- 10R(iv) ---
                            elif sheet_name == "10R(iv)":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                b_col = df_raw.iloc[:, 1].astype(str).str.strip()
                                
                                # --- Table 1 Logic (A11 -> Unsecured loans - 1) ---
                                t1_end_idx = next((i for i, v in enumerate(b_col) if "Unsecured loans" in v), None)
                                if t1_end_idx is not None:
                                    df_t1 = df_raw.iloc[10:t1_end_idx, 0:5].copy() # A:E
                                    df_t1 = df_t1[df_t1.iloc[:, 1].notna() & (df_t1.iloc[:, 1].astype(str).str.strip() != "")]
                                    if not df_t1.empty:
                                        df_t1.iloc[:, 0] = str(ac_value) # A-г Компани кодоор солих
                                        all_10riv_t1_dfs.append(df_t1)

                                # --- Table 2 Logic (t1_end + 1 -> Others - 1) ---
                                # startTable2Range = endTable1Row + 2. Индексээр бол t1_end_idx + 1
                                t2_start_idx = t1_end_idx + 1 if t1_end_idx is not None else 17
                                t2_end_idx = next((i for i, v in enumerate(b_col) if "Others" in v), None)
                                if t2_end_idx is not None:
                                    df_t2 = df_raw.iloc[t2_start_idx:t2_end_idx, 0:5].copy()
                                    df_t2 = df_t2[df_t2.iloc[:, 1].notna() & (df_t2.iloc[:, 1].astype(str).str.strip() != "")]
                                    if not df_t2.empty:
                                        df_t2.iloc[:, 0] = str(ac_value)
                                        all_10riv_t2_dfs.append(df_t2)

                                # --- Table 3 Logic (t2_end + 1 -> Total interest-bearing borrowings - 1) ---
                                t3_start_idx = t2_end_idx + 1 if t2_end_idx is not None else 23
                                t3_end_idx = next((i for i, v in enumerate(b_col) if "Total interest-bearing borrowings" in v), None)
                                if t3_end_idx is not None:
                                    df_t3 = df_raw.iloc[t3_start_idx:t3_end_idx, 0:5].copy()
                                    df_t3 = df_t3[df_t3.iloc[:, 1].notna() & (df_t3.iloc[:, 1].astype(str).str.strip() != "")]
                                    if not df_t3.empty:
                                        df_t3.iloc[:, 0] = str(ac_value)
                                        all_10riv_t3_dfs.append(df_t3)

                            # --- 10R(v) ---
                            elif sheet_name == "10R(v)":
                                # F4:F21 хүрээг унших (Python индексээр Row 3:21, Col 5)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # F4:F21 датаг авах (3-р мөрнөөс 21-р мөр хүртэл, 5-р багана)
                                # pandas.iloc[3:21, 5] гэдэг нь Excel-ийн F4:F21 юм
                                r5_col_data = df_raw.iloc[3:21, 5].tolist()
                                
                                if r5_col_data:
                                    # F4 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r5_col_data[0] = str(ac_value)
                                    all_10rv_cols.append(r5_col_data)

                            # --- 11det1 ---
                            elif sheet_name == "11det1":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. "ЗААВАР:" текстээр төгсгөлийн мөрийг олох
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                end_idx = next((i for i, v in enumerate(a_col) if "ЗААВАР:" in v), None)
                                
                                if end_idx is not None:
                                    # Төгсгөлийн мөр: end_idx - 9 (Таны заавраар)
                                    # А9-өөс эхлэх тул эхлэх индекс 8
                                    read_end_row = end_idx - 9
                                    df_part = df_raw.iloc[8:read_end_row, :].copy()
                                    
                                    # 2. Зөвхөн шаардлагатай багануудыг сонгох (Excel Column -> Python Index)
                                    # A:R(0:18), T:X(19:24), Z:AA(25:27), AC:AD(28:30), AF:AG(31:33), AI:AJ(34:36), AL:AM(37:39)
                                    target_cols = list(range(0, 18)) + [19, 20, 21, 22, 23] + [25, 26] + \
                                                  [28, 29] + [31, 32] + [34, 35] + [37, 38]
                                    
                                    # Баганууд байгаа эсэхийг баталгаажуулж reindex хийх
                                    max_col_needed = max(target_cols)
                                    if df_part.shape[1] <= max_col_needed:
                                        df_part = df_part.reindex(columns=range(max_col_needed + 1))
                                    
                                    df_part = df_part.iloc[:, target_cols]
                                    
                                    # 3. ШҮҮЛТҮҮР: F багана (Index 5) хоосон байвал тухайн row-ийг устгах
                                    # Анхаар: target_cols-д F багана (5) орсон байгаа тул индекс хэвээрээ байна
                                    df_part = df_part[df_part.iloc[:, 5].notna()]
                                    df_part = df_part[df_part.iloc[:, 5].astype(str).str.strip() != ""]
                                    
                                    if not df_part.empty:
                                        # 4. A баганыг (Index 0) компани кодоор солих
                                        df_part.iloc[:, 0] = str(ac_value)
                                        all_11det1_dfs.append(df_part)
                            
                            # --- 11det2 ---
                            elif sheet_name == "11det2":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. "ЗААВАР:" текстээр төгсгөлийн мөрийг олох
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                end_idx = next((i for i, v in enumerate(a_col) if "ЗААВАР:" in v), None)
                                
                                if end_idx is not None:
                                    # Төгсгөлийн мөр: end_idx - 24 (Таны заавраар)
                                    # A12-оос эхлэх тул эхлэх индекс 11
                                    read_end_row = end_idx - 24
                                    df_part = df_raw.iloc[11:read_end_row, :].copy()
                                    
                                    # 2. Зөвхөн шаардлагатай багануудыг сонгох (Excel Column -> Python Index)
                                    # A:B(0,1), F:J(5:10), M:N(12:14), Q:AB(16:28), AH:AM(33:39), AO:AT(40:46)
                                    target_cols = [0, 1] + list(range(5, 10)) + [12, 13] + \
                                                  list(range(16, 28)) + list(range(33, 39)) + \
                                                  list(range(40, 46))
                                    
                                    # Хэрэв Excel файл AT (45) багана хүртэл хүрэхгүй бол reindex хийж нөхөх
                                    max_col_needed = max(target_cols)
                                    if df_part.shape[1] <= max_col_needed:
                                        df_part = df_part.reindex(columns=range(max_col_needed + 1))
                                    
                                    df_part = df_part.iloc[:, target_cols]
                                    
                                    # 3. ШҮҮЛТҮҮР: B багана (Index 1) хоосон байвал тухайн row-ийг устгах
                                    # Манай df_part-ийн 2 дахь багана (index 1) бол Excel-ийн B багана юм
                                    df_part = df_part[df_part.iloc[:, 1].notna()]
                                    df_part = df_part[df_part.iloc[:, 1].astype(str).str.strip() != ""]
                                    
                                    if not df_part.empty:
                                        # 4. A баганыг (Index 0) компани кодоор солих
                                        df_part.iloc[:, 0] = str(ac_value)
                                        all_11det2_dfs.append(df_part)

                            # --- 11R ---
                            elif sheet_name == "11R":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # F:F (5), J:J (9), N:N (13) багануудын 3-аас 22-р мөрийг унших
                                # Python индексээр 2:22 (3-р мөрөөс 22 хүртэл)
                                f_col = df_raw.iloc[2:22, 5].tolist()
                                j_col = df_raw.iloc[2:22, 9].tolist()
                                n_col = df_raw.iloc[2:22, 13].tolist()
                                
                                # Компани бүрийн 4-р мөрийг (жагсаалтын 1 дэх индекс) кодоор солих
                                if len(f_col) > 1: f_col[1] = str(ac_value)
                                if len(j_col) > 1: j_col[1] = str(ac_value)
                                if len(n_col) > 1: n_col[1] = str(ac_value)
                                
                                # 3 баганаа нэг багц болгож хадгалах
                                all_11r_cols.append([f_col, j_col, n_col])

                            # --- 11R(i) ---
                            elif sheet_name == "11R(i)":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # E:E (4), G:G (6) багануудын 4-өөс 23-р мөрийг унших
                                # Python индексээр 3:23 (4-р мөрөөс 23 хүртэл)
                                e_col = df_raw.iloc[3:23, 4].tolist()
                                g_col = df_raw.iloc[3:23, 6].tolist()
                                
                                # E4 болон G4 нүднүүдийг (жагсаалтын эхний элемент) кодоор солих
                                if e_col: e_col[0] = str(ac_value)
                                if g_col: g_col[0] = str(ac_value)
                                
                                # 2 баганаа нэг багц болгож хадгалах
                                all_11ri_cols.append([e_col, g_col])

                            # --- 12det ---
                            elif sheet_name == "12det":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. "ЗААВАР:" текстээр төгсгөлийн мөрийг олох
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                end_idx = next((i for i, v in enumerate(a_col) if "ЗААВАР:" in v), None)
                                
                                if end_idx is not None:
                                    # Төгсгөлийн мөр: end_idx - 24, Эхлэх индекс 11 (A12)
                                    read_end_row = end_idx - 24
                                    df_part = df_raw.iloc[11:read_end_row, :].copy()
                                    
                                    # 2. Шаардлагатай багануудыг тодорхойлох (AH:AN нэмэгдсэн)
                                    # A:B(0,1), G:J(6:9), L(11), R:W(17:22), Z:AA(25:26), AH:AN(33:39)
                                    # Тэмдэглэл: range(start, end) функц end-ийг оруулахгүй тул +1 авна
                                    target_cols = [0, 1] + list(range(6, 10)) + [11] + \
                                                  list(range(17, 23)) + [25, 27] + \
                                                  list(range(33, 40)) 
                                    
                                    # 3. Алдаанаас сэргийлж баганын тоог нөхөх (AN хүртэл буюу 39 индекс)
                                    max_col_needed = max(target_cols)
                                    if df_part.shape[1] <= max_col_needed:
                                        # Баганын тоо дутуу бол None-оор дүүргэсэн шинэ баганууд нэмнэ
                                        new_cols = range(df_part.shape[1], max_col_needed + 1)
                                        for c in new_cols:
                                            df_part[c] = None
                                    
                                    # Зөвхөн хэрэгцээт багануудыг салгаж авах
                                    df_part = df_part.iloc[:, target_cols]
                                    
                                    # 4. ШҮҮЛТҮҮР: B багана (Index 1) хоосон байвал тухайн row-ийг устгах
                                    df_part = df_part[df_part.iloc[:, 1].notna()]
                                    df_part = df_part[df_part.iloc[:, 1].astype(str).str.strip() != ""]
                                    
                                    if not df_part.empty:
                                        # 5. A баганыг (Index 0) компани кодоор солих
                                        df_part.iloc[:, 0] = str(ac_value)
                                        all_12det_dfs.append(df_part)

                            # --- 12R ---
                            elif sheet_name == "12R":
                                # F2:F45 хүрээг унших (Python индексээр Row 1:45, Col 5)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # F2:F45 датаг авах (2-р мөрөөс 45-р мөр хүртэл, 5-р багана)
                                # pandas.iloc[1:45, 5] гэдэг нь Excel-ийн F2:F45 юм
                                r12_col_data = df_raw.iloc[1:45, 5].tolist()
                                
                                if r12_col_data:
                                    # F2 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r12_col_data[0] = str(ac_value)
                                    all_12r_cols.append(r12_col_data)

                            # 13det
                            elif sheet_name == "13det":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                
                                # --- Table 1-ийг салгаж авах ---
                                t1_text = "B. Provision - Current (313001, 313002, 313003, 313004, 313092, 312597)"
                                t1_end_idx = next((i for i, v in enumerate(a_col) if t1_text in v), None)
                                
                                if t1_end_idx is not None:
                                    t1_read_end = t1_end_idx - 4
                                    # Excel A7-оос тэрхүү текст хүртэл (index 6:t1_read_end)
                                    df_t1 = df_raw.iloc[6:t1_read_end, :].copy()
                                    
                                    target_cols = [0, 1, 2] + list(range(4, 12)) # A:C, E:L
                                    df_t1 = df_t1.reindex(columns=range(12))
                                    df_t1 = df_t1.iloc[:, target_cols]
                                    # C багана (Index 2) хоосон биш байх
                                    df_t1 = df_t1[df_t1.iloc[:, 2].notna() & (df_t1.iloc[:, 2].astype(str).str.strip() != "")]
                                    
                                    if not df_t1.empty:
                                        df_t1.iloc[:, 0] = str(ac_value)
                                        all_13det_t1_dfs.append(df_t1)
                                    
                                    # --- Table 2-ийг салгаж авах ---
                                    # Файл дээрх Table 2-ийн эхлэл = Table 1-ийн төгсгөл + 7 мөр
                                    t2_file_start = t1_read_end + 8 
                                    t2_text = "Provision category:"
                                    t2_end_idx = next((i for i, v in enumerate(a_col) if t2_text in v), None)
                                    
                                    if t2_end_idx is not None:
                                        t2_read_end = t2_end_idx - 5
                                        # Файл бүрийн өөр өөр байрлалаас Table 2-ийг уншина
                                        df_t2 = df_raw.iloc[t2_file_start-1:t2_read_end, :].copy()
                                        
                                        df_t2 = df_t2.reindex(columns=range(12))
                                        df_t2 = df_t2.iloc[:, target_cols]
                                        # C багана (Index 2) хоосон биш байх
                                        df_t2 = df_t2[df_t2.iloc[:, 2].notna() & (df_t2.iloc[:, 2].astype(str).str.strip() != "")]
                                        
                                        if not df_t2.empty:
                                            df_t2.iloc[:, 0] = str(ac_value)
                                            all_13det_t2_dfs.append(df_t2)

                            # --- 13R ---
                            elif sheet_name == "13R":
                                # D4:D41 хүрээг унших (Python индексээр Row 3:41, Col 3)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # D4:D41 датаг авах (3-р мөрнөөс 41-р мөр хүртэл, 3-р багана)
                                # pandas.iloc[3:41, 3] гэдэг нь Excel-ийн D4:D41 юм
                                r13_col_data = df_raw.iloc[3:41, 3].tolist()
                                
                                if r13_col_data:
                                    # D4 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r13_col_data[0] = str(ac_value)
                                    all_13r_cols.append(r13_col_data)

                            # --- 13R(i) ---
                            elif sheet_name == "13R(i)":
                                # F4:F19 хүрээг унших (Python индексээр Row 3:19, Col 5)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # F4:F19 датаг авах (3-р мөрнөөс 19-р мөр хүртэл, 5-р багана)
                                # pandas.iloc[3:19, 5] гэдэг нь Excel-ийн F4:F19 юм
                                r13i_col_data = df_raw.iloc[3:19, 5].tolist()
                                
                                if r13i_col_data:
                                    # F4 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r13i_col_data[0] = str(ac_value)
                                    all_13ri_cols.append(r13i_col_data)

                            # --- 14R ---
                            elif sheet_name == "14R":
                                # E3:E21 хүрээг унших (Python индексээр Row 2:21, Col 4)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # E3:E21 датаг авах (2-р мөрөөс 21-р мөр хүртэл, 4-р багана)
                                # pandas.iloc[2:21, 4] гэдэг нь Excel-ийн E3:E21 юм
                                r14_col_data = df_raw.iloc[2:21, 4].tolist()
                                
                                if r14_col_data:
                                    # E3 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r14_col_data[0] = str(ac_value)
                                    all_14r_cols.append(r14_col_data)

                            # --- 14R(i) ---
                            elif sheet_name.strip() == "14R(i)":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                
                                b_idx = next((i for i, v in enumerate(a_col) if "B. Advances from customers: 321000-321003; 351000-351003" in v), None)
                                c_idx = next((i for i, v in enumerate(a_col) if "ЗААВАР:" in v), None)
                                
                                if b_idx is not None and c_idx is not None:
                                    a_idx = next((i for i in range(5, b_idx) if "Non-current" in str(df_raw.iloc[i, 0])), None)
                                    d_idx = next((i for i in range(b_idx, c_idx) if "Non-current" in str(df_raw.iloc[i, 0])), None)
                                    
                                    def process_table(start_r, end_r):
                                        t_df = df_raw.iloc[start_r:end_r, 0:12].copy() # A:L унших
                                        t_df = t_df[t_df.iloc[:, 4].notna()] # E column-оор шүүх
                                        t_df = t_df[pd.to_numeric(t_df.iloc[:, 4], errors='coerce').fillna(0) != 0]
                                        
                                        if not t_df.empty:
                                            # !!! ЧУХАЛ: Энд A баганыг өөрчлөхгүй !!!
                                            # Харин (DataFrame, CompanyCode) гэсэн багц буцаана
                                            return (t_df, str(ac_value))
                                        return None

                                    # Table 1
                                    if a_idx:
                                        res = process_table(5, a_idx - 2)
                                        if res: all_14r1_t1.append(res)
                                        # Table 2
                                        res = process_table(a_idx + 1, b_idx - 5)
                                        if res: all_14r1_t2.append(res)
                                    # Table 3
                                    if d_idx:
                                        res = process_table(b_idx + 4, d_idx - 2)
                                        if res: all_14r1_t3.append(res)
                                        # Table 4
                                        res = process_table(d_idx + 1, c_idx - 6)
                                        if res: all_14r1_t4.append(res)

                            # --- 17det1 ---
                            elif sheet_name == "17det1":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. "ЗААВАР:" текстээр төгсгөлийн мөрийг олох
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                end_idx = next((i for i, v in enumerate(a_col) if "ЗААВАР:" in v), None)
                                
                                if end_idx is not None:
                                    # Төгсгөлийн мөр: end_idx - 3
                                    # A4-өөс эхлэх тул эхлэх индекс 3
                                    read_end_row = end_idx - 3
                                    df_part = df_raw.iloc[3:read_end_row, :].copy()
                                    
                                    # 2. Зөвхөн шаардлагатай багануудыг сонгох (Excel Column -> Python Index)
                                    # A:C (0,1,2) болон E:L (4,5,6,7,8,9,10,11)
                                    target_cols = [0, 1, 2] + list(range(4, 12))
                                    
                                    # Баганын тоог нөхөж алдаанаас сэргийлэх (L багана хүртэл буюу 11 индекс)
                                    max_col_needed = max(target_cols)
                                    if df_part.shape[1] <= max_col_needed:
                                        df_part = df_part.reindex(columns=range(max_col_needed + 1))
                                    
                                    df_part = df_part.iloc[:, target_cols]
                                    
                                    # 3. ШҮҮЛТҮҮР: B багана (Index 1) хоосон байвал тухайн row-ийг устгах
                                    df_part = df_part[df_part.iloc[:, 1].notna()]
                                    df_part = df_part[df_part.iloc[:, 1].astype(str).str.strip() != ""]
                                    
                                    if not df_part.empty:
                                        # 4. A баганыг (Index 0) компани кодоор солих
                                        df_part.iloc[:, 0] = str(ac_value)
                                        all_17det1_dfs.append(df_part)

                            # --- 17det2 ---
                            elif sheet_name == "17det2":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. "Type of customers" текстээр төгсгөлийн мөрийг олох
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                end_idx = next((i for i, v in enumerate(a_col) if "Type of customers" in v), None)
                                
                                if end_idx is not None:
                                    # Төгсгөлийн мөр: end_idx - 3
                                    # A7-оос эхлэх тул эхлэх индекс 6
                                    read_end_row = end_idx - 3
                                    df_part = df_raw.iloc[6:read_end_row, :].copy()
                                    
                                    # 2. Зөвхөн шаардлагатай багануудыг сонгох (Excel Column -> Python Index)
                                    # A:F (0:6), H:L (7:12), O:T (14:20)
                                    target_cols = list(range(0, 6)) + list(range(7, 12)) + list(range(14, 20))
                                    
                                    # Баганын тоог нөхөж алдаанаас сэргийлэх (T багана хүртэл буюу 19 индекс)
                                    max_col_needed = max(target_cols)
                                    if df_part.shape[1] <= max_col_needed:
                                        df_part = df_part.reindex(columns=range(max_col_needed + 1))
                                    
                                    df_part = df_part.iloc[:, target_cols]
                                    
                                    # 3. ШҮҮЛТҮҮР: B багана (Index 1) хоосон байвал тухайн row-ийг устгах
                                    df_part = df_part[df_part.iloc[:, 1].notna()]
                                    df_part = df_part[df_part.iloc[:, 1].astype(str).str.strip() != ""]
                                    
                                    if not df_part.empty:
                                        # 4. A баганыг (Index 0) компани кодоор солих
                                        df_part.iloc[:, 0] = str(ac_value)
                                        all_17det2_dfs.append(df_part)

                            # --- 17R(ii) ---
                            elif sheet_name.strip() == "17R(ii)":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # D2:D30 (Row 1:30, Col 3), F2:F30 (Col 5), H2:H30 (Col 7)
                                d_col = df_raw.iloc[1:30, 3].tolist()
                                f_col = df_raw.iloc[1:30, 5].tolist()
                                h_col = df_raw.iloc[1:30, 7].tolist()
                                
                                if d_col or f_col or h_col:
                                    # 3 баганыг нэг багц (tuple) болгож компани кодын хамт хадгалах
                                    all_17rii_data.append({
                                        'code': str(ac_value),
                                        'columns': [d_col, f_col, h_col]
                                    })

                            # ICT Report Raw Summary
                            elif sheet_name == "ICT Report Raw Summary":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. Дата байгаа эсэхийг шалгах (Row 7, Col F буюу индекс [6, 5])
                                # Хэрэв F7-д утга байхгүй байсан ч F8, F9-д байж болох тул 
                                # F7-оос доошхи хэсгийг бүхэлд нь шалгая
                                ict_data_area = df_raw.iloc[6:, 5:19] # F7:S хүртэлх хэсэг
                                
                                # Хэрэв энэ хэсэгт ямар нэг өгөгдөл байвал уншина
                                if not ict_data_area.dropna(how='all').empty:
                                    # F6-аас эхлэн (индекс 5) S багана (индекс 18) хүртэл унших
                                    # last_valid_index-ээс илүү найдвартай арга:
                                    # Дата байгаа хамгийн сүүлийн мөрийг ict_data_area-аас олох
                                    mask = df_raw.iloc[:, 5:19].notna().any(axis=1)
                                    last_row_idx = mask[mask].index[-1]
                                    
                                    # F6-аас (Index 5) сүүлийн мөр хүртэл унших
                                    df_ict = df_raw.iloc[5:last_row_idx + 1, 5:19].copy()
                                    
                                    # Шинэ T багана үүсгэж компанийн код бичих
                                    df_ict['company_code'] = str(ac_value)
                                    
                                    all_ict_summary_dfs.append(df_ict)
                    
                            # --- 17R ---
                            elif sheet_name == "17R":
                                # E3:E64 хүрээг унших (Python индексээр Row 2:64, Col 4)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # E3:E64 датаг авах (2-р мөрөөс 64-р мөр хүртэл, 4-р багана)
                                # pandas.iloc[2:64, 4] гэдэг нь Excel-ийн E3:E64 юм
                                r17_col_data = df_raw.iloc[2:64, 4].tolist()
                                
                                if r17_col_data:
                                    # E3 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r17_col_data[0] = str(ac_value)
                                    all_17r_cols.append(r17_col_data)

                            # --- 18det ---
                            elif sheet_name.strip() == "18det":
                                # G1:G101 хүрээг унших (Python индексээр Row 0:101, Col 6)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # G1:G101 датаг авах (1-р мөрөөс 101-р мөр хүртэл, 6-р багана)
                                # pandas.iloc[0:101, 6] гэдэг нь Excel-ийн G1:G101 юм
                                r18_col_data = df_raw.iloc[0:101, 6].tolist()
                                
                                if r18_col_data:
                                    # G1 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r18_col_data[0] = str(ac_value)
                                    all_18det_cols.append(r18_col_data)

                            # --- 19R ---
                            elif sheet_name.strip() == "19R":
                                # E4:E37 хүрээг унших (Python индексээр Row 3:37, Col 4)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # E4:E37 датаг авах (4-р мөрөөс 37-р мөр хүртэл, 4-р багана)
                                # pandas.iloc[3:37, 4] гэдэг нь Excel-ийн E4:E37 юм
                                r19_col_data = df_raw.iloc[3:37, 4].tolist()
                                
                                if r19_col_data:
                                    # E4 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r19_col_data[0] = str(ac_value)
                                    all_19r_cols.append(r19_col_data)

                            # --- 20R ---
                            elif sheet_name.strip() == "20R":
                                # F3:F69 хүрээг унших (Python индексээр Row 2:69, Col 5)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # F3:F69 датаг авах (3-р мөрөөс 69-р мөр хүртэл, 5-р багана)
                                # pandas.iloc[2:69, 5] гэдэг нь Excel-ийн F3:F69 юм
                                r20_col_data = df_raw.iloc[2:69, 5].tolist()
                                
                                if r20_col_data:
                                    # F3 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r20_col_data[0] = str(ac_value)
                                    all_20r_cols.append(r20_col_data)
                            
                            # --- 20R(i) ---
                            elif sheet_name.strip() == "20R(i)":
                                # D3:D31 хүрээг унших (Python индексээр Row 2:31, Col 3)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # D3:D31 датаг авах (3-р мөрөөс 31-р мөр хүртэл, 3-р багана)
                                # pandas.iloc[2:31, 3] гэдэг нь Excel-ийн D3:D31 юм
                                r20i_col_data = df_raw.iloc[2:31, 3].tolist()
                                
                                if r20i_col_data:
                                    # D3 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r20i_col_data[0] = str(ac_value)
                                    all_20ri_cols.append(r20i_col_data)

                            # --- 21R ---
                            elif sheet_name.strip() == "21R":
                                # E3:E65 хүрээг унших (Python индексээр Row 2:65, Col 4)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # E3:E65 датаг авах (3-р мөрөөс 65-р мөр хүртэл, 4-р багана)
                                # pandas.iloc[2:65, 4] гэдэг нь Excel-ийн E3:E65 юм
                                r21_col_data = df_raw.iloc[2:65, 4].tolist()
                                
                                if r21_col_data:
                                    # E3 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r21_col_data[0] = str(ac_value)
                                    all_21r_cols.append(r21_col_data)

                            # --- 21R(i) ---
                            elif sheet_name.strip() == "21R(i)":
                                # C4:C41 хүрээг унших (Python индексээр Row 3:41, Col 2)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # C4:C41 датаг авах (4-р мөрөөс 41-р мөр хүртэл, 3-р багана)
                                # pandas.iloc[3:41, 2] гэдэг нь Excel-ийн C4:C41 юм
                                r21i_col_data = df_raw.iloc[3:41, 2].tolist()
                                
                                if r21i_col_data:
                                    # C4 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r21i_col_data[0] = str(ac_value)
                                    all_21ri_cols.append(r21i_col_data)

                            # --- 22R(i) ---
                            elif sheet_name.strip() == "22R(i)":
                                # F5:F13 хүрээг унших (Python индексээр Row 4:13, Col 5)
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # F5:F13 датаг авах (5-р мөрөөс 13-р мөр хүртэл, 6-р багана)
                                # pandas.iloc[4:13, 5] гэдэг нь Excel-ийн F5:F13 юм
                                r22i_col_data = df_raw.iloc[4:13, 5].tolist()
                                
                                if r22i_col_data:
                                    # F5 нүдийг (жагсаалтын эхний элемент) компани кодоор солих
                                    r22i_col_data[0] = str(ac_value)
                                    all_22ri_cols.append(r22i_col_data)

                            # --- 22R(vi) ---
                            elif sheet_name.strip() == "22R(vi)":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. F6:F32 унших (Index: Row 5:32, Col 5)
                                f_col_data = df_raw.iloc[5:32, 5].tolist()
                                # 2. H6:H32 унших (Index: Row 5:32, Col 7)
                                h_col_data = df_raw.iloc[5:32, 7].tolist()
                                
                                if f_col_data and h_col_data:
                                    # F6 нүдийг (F жагсаалтын эхний элемент) компани кодоор солих
                                    f_col_data[0] = str(ac_value)
                                    # Хоёр баганыг нэг компани доор багцалж хадгалах
                                    all_22rvi_cols.append((f_col_data, h_col_data))

                            # --- 24det ---
                            elif sheet_name.strip() == "24det":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. "ЗААВАР:" текстээр төгсгөлийн мөрийг олох
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                end_idx = next((i for i, v in enumerate(a_col) if "ЗААВАР:" in v), None)
                                
                                if end_idx is not None:
                                    # Төгсгөлийн мөр: end_idx - 22
                                    # A7-оос эхлэх тул эхлэх индекс 6
                                    read_end_row = end_idx - 22
                                    df_part = df_raw.iloc[6:read_end_row, :].copy()
                                    
                                    # 2. Зөвхөн шаардлагатай багануудыг сонгох A:L (Index 0-оос 11)
                                    df_part = df_part.reindex(columns=range(12))
                                    df_part = df_part.iloc[:, 0:12]
                                    
                                    # 3. ШҮҮЛТҮҮР: B багана (Index 1) хоосон байвал тухайн row-ийг устгах
                                    df_part = df_part[df_part.iloc[:, 1].notna()]
                                    df_part = df_part[df_part.iloc[:, 1].astype(str).str.strip() != ""]
                                    
                                    if not df_part.empty:
                                        # 4. A баганыг (Index 0) компани кодоор солих
                                        df_part.iloc[:, 0] = str(ac_value)
                                        all_24det_dfs.append(df_part)

                            # --- 24R(i) ---
                            elif sheet_name.strip() == "24R(i)":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # F4:F20 хүрээг унших
                                raw_values = df_raw.iloc[3:20, 5].tolist()
                                
                                # А. Нийлбэр бодоход зориулж: Тоо руу хөрвүүлэх (текст/хоосныг 0 болгоно)
                                numeric_for_sum = pd.to_numeric(raw_values, errors='coerce')
                                sum_ready_list = [float(v) if pd.notna(v) else 0.0 for v in numeric_for_sum]
                                all_24ri_sum_data.append(sum_ready_list)
                                
                                # Б. Бичихэд зориулж: Анхны утгыг хадгалах
                                col_to_move = list(raw_values)
                                col_to_move[0] = str(ac_value) # F4-ийг компани кодоор солих
                                all_24ri_cols.append(col_to_move)

                            # --- 24R(ii) ---
                            elif sheet_name.strip() == "24R(ii)":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # F5:F11 хүрээг унших (Python индексээр Row 4:11, Col 5)
                                raw_values_24rii = df_raw.iloc[4:11, 5].tolist()
                                
                                if raw_values_24rii:
                                    # Хуулбар авч, эхний нүдийг (F5) компани кодоор солих
                                    col_to_move = list(raw_values_24rii)
                                    col_to_move[0] = str(ac_value)
                                    all_24rii_cols.append(col_to_move)

                            # --- 25det1 ---
                            elif sheet_name.strip() == "25det1":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. Төгсгөлийн мөрийг олох (ЗААВАР: - 3)
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                end_idx = next((i for i, v in enumerate(a_col) if "ЗААВАР:" in v), None)
                                
                                if end_idx is not None:
                                    read_end_row = end_idx - 3
                                    # A7-оос эхлэх тул индекс 6-аас read_end_row хүртэл
                                    df_part = df_raw.iloc[6:read_end_row, :].copy()
                                    
                                    # 2. ШҮҮЛТҮҮР: B багана (Index 1) хоосон байвал мөрийг устгах
                                    df_part = df_part[df_part.iloc[:, 1].notna()]
                                    df_part = df_part[df_part.iloc[:, 1].astype(str).str.strip() != ""]
                                    
                                    if not df_part.empty:
                                        # 3. Зөвхөн заасан багануудыг сонгож авах
                                        # A=0, B=1, D=3, E=4, G=6, J=9, K=10, L=11
                                        cols_to_keep = [0, 1, 3, 4, 6, 9, 10, 11]
                                        df_filtered = df_part.iloc[:, cols_to_keep].copy()
                                        
                                        # 4. A баганыг (Шүүгдсэн датаны 0-р багана) компани кодоор солих
                                        df_filtered.iloc[:, 0] = str(ac_value)
                                        
                                        all_25det1_dfs.append(df_filtered)

                            # --- 25det2 ---
                            elif sheet_name.strip() == "25det2":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. Төгсгөлийн мөрийг олох (ЗААВАР: - 3)
                                a_col = df_raw.iloc[:, 0].astype(str).str.strip()
                                end_idx = next((i for i, v in enumerate(a_col) if "ЗААВАР:" in v), None)
                                
                                if end_idx is not None:
                                    read_end_row = end_idx - 3
                                    # A7-оос эхлэх тул индекс 6
                                    df_part = df_raw.iloc[6:read_end_row, :].copy()
                                    
                                    # 2. ШҮҮЛТҮҮР: B багана (Index 1) хоосон байвал мөрийг устгах
                                    df_part = df_part[df_part.iloc[:, 1].notna()]
                                    df_part = df_part[df_part.iloc[:, 1].astype(str).str.strip() != ""]
                                    
                                    if not df_part.empty:
                                        # 3. Зөвхөн заасан багануудыг сонгож авах
                                        # A=0, B=1, D=3, F=5, G=6
                                        cols_to_keep = [0, 1, 3, 5, 6]
                                        df_filtered = df_part.iloc[:, cols_to_keep].copy()
                                        
                                        # 4. A баганыг компанийн кодоор солих
                                        df_filtered.iloc[:, 0] = str(ac_value)
                                        
                                        all_25det2_dfs.append(df_filtered)

                            # --- 25R(i) ---
                            elif sheet_name.strip() == "25R(i)":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # L5:L41 хүрээг унших (Python индексээр Row 4:41, Col 11)
                                raw_values_25ri = df_raw.iloc[4:41, 11].tolist()
                                
                                if raw_values_25ri:
                                    # Хуулбар авч, эхний нүдийг (L5) компани кодоор солих
                                    col_to_move = list(raw_values_25ri)
                                    col_to_move[0] = str(ac_value)
                                    all_25ri_cols.append(col_to_move)

                            # --- 25R(ii) ---
                            elif sheet_name.strip() == "25R(ii)":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. F4:F15 хүрээг унших (Индекс 3-аас 15 хүртэл)
                                raw_values_25rii = df_raw.iloc[3:15, 5].tolist()
                                
                                # 2. Нийлбэр бодоход зориулж: Тоо руу хөрвүүлэх (текст/хоосныг 0 болгоно)
                                numeric_for_sum = pd.to_numeric(raw_values_25rii, errors='coerce')
                                sum_ready_list = [float(v) if pd.notna(v) else 0.0 for v in numeric_for_sum]
                                all_25rii_sum_data.append(sum_ready_list)
                                
                                # 3. Бичихэд зориулж: Анхны утгыг хадгалах
                                col_to_move = list(raw_values_25rii)
                                col_to_move[0] = str(ac_value) # F4-ийг компани кодоор солих
                                all_25rii_cols.append(col_to_move)

                            # --- Sheet 26 ---
                            elif sheet_name.strip() == "26":
                                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=object)
                                
                                # 1. C2:C10 хүрээг унших (Python индексээр Row 1:10, Col 2)
                                raw_values_26 = df_raw.iloc[1:10, 2].tolist()
                                
                                # 2. Тоо руу хөрвүүлэх (Нийлбэр болон 0-ээр орлуулахад зориулж)
                                numeric_26 = pd.to_numeric(raw_values_26, errors='coerce')
                                # Хоосон (NaN) бол 0.0 болгох
                                processed_values = [float(v) if pd.notna(v) else 0.0 for v in numeric_26]
                                
                                if processed_values:
                                    all_26_sum_data.append(processed_values)
                                    
                                    # 3. Компани бүрээр цувуулах дата бэлдэх
                                    col_to_move = list(processed_values)
                                    col_to_move[0] = str(ac_value) # C2-ийг компани кодоор солих
                                    all_26_cols.append(col_to_move)

                        except Exception as e:
                            # Алдааг жагсаалтад бүртгэнэ
                            err_msg = f"❌ {file_name} -> {sheet_name} хуудас дээр алдаа: {str(e)}"
                            error_logs.append(err_msg)
                            # Процессыг зогсоохгүй, дараагийн файл руу шилжүүлнэ
                            continue

# --- Template-д бичих ---
                    if sheet_name in book.sheetnames:
                        ts = book[sheet_name]
                        
                        if sheet_name == "1.1det" and generic_dfs:
                            f = pd.concat(generic_dfs, ignore_index=True)
                            skip_idx = [11, 17, 18, 19, 22] # L, R, S, T, W
                            for r_i, row in enumerate(f.values, start=8):
                                for c_i, val in enumerate(row):
                                    if c_i not in skip_idx:
                                        if pd.notna(val): ts.cell(row=r_i, column=c_i+1).value = val
                        
                        elif sheet_name == "1.2det" and generic_dfs:
                            f = pd.concat(generic_dfs, ignore_index=True)
                            for r_i, row in enumerate(f.values, start=7):
                                for c_i, val in enumerate(row, start=1):
                                    if pd.notna(val): ts.cell(row=r_i, column=c_i).value = val

                        elif sheet_name == "1.3det":
                            if t1_13_dfs:
                                f1 = pd.concat(t1_13_dfs, ignore_index=True)
                                for r_i, row in enumerate(f1.values, start=8):
                                    for c_i, val in enumerate(row, start=1):
                                        if pd.notna(val): ts.cell(row=r_i, column=c_i).value = val
                            if t2_13_dfs:
                                f2 = pd.concat(t2_13_dfs, ignore_index=True)
                                for r_i, row in enumerate(f2.values, start=61):
                                    for c_i, val in enumerate(row, start=1):
                                        if pd.notna(val): ts.cell(row=r_i, column=c_i).value = val

                        elif sheet_name == "1R(i)":
                            ts["F12"] = r1_sum_val
                            for c_i, cv in enumerate(r1_i_cols):
                                for r_i, val in enumerate(cv, start=4):
                                    if pd.notna(val): ts.cell(row=r_i, column=11+c_i).value = val

                        elif sheet_name == "1R(ii)":
                            for c_i, cv in enumerate(r1_ii_cols):
                                for r_i, val in enumerate(cv, start=5):
                                    if pd.notna(val): ts.cell(row=r_i, column=12+c_i).value = val

                        elif sheet_name == "1R(iii)":
                            for c_i, cv in enumerate(r1_iii_cols):
                                for r_i, val in enumerate(cv, start=3):
                                    if pd.notna(val): ts.cell(row=r_i, column=12+c_i).value = val

                        elif sheet_name == "1R(iv)":
                            for c_i, cv in enumerate(r1_iv_cols):
                                for r_i, val in enumerate(cv, start=4):
                                    if pd.notna(val): ts.cell(row=r_i, column=12+c_i).value = val

                        elif sheet_name == "2R":
                            # Бичих багануудын Excel дугаар: A-K (1-11), O (15), P (16)
                            cols_2r = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 15, 16]
                            
                            if r2_dfs:
                                f2r = pd.concat(r2_dfs, ignore_index=True)
                                # A7 мөрнөөс эхэлж бичих
                                for r_i, row in enumerate(f2r.values, start=7):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=r_i, column=cols_2r[c_idx]).value = val

                        elif sheet_name == "3R":
                            # Бичих багануудын Excel дугаар: A-M (1-13), O,P (15,16), R,S (18,19), U,V (21,22), Y,Z (25,26)
                            cols_3r = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 18, 19, 21, 22, 25, 26]
                            
                            # Table 1-ийг A8-аас эхэлж бичих
                            if t1_3r_dfs:
                                f1 = pd.concat(t1_3r_dfs, ignore_index=True)
                                for r_i, row in enumerate(f1.values, start=8):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=r_i, column=cols_3r[c_idx]).value = val
                                            
                            # Table 2-ийг A86-аас эхэлж бичих
                            if t2_3r_dfs:
                                f2 = pd.concat(t2_3r_dfs, ignore_index=True)
                                for r_i, row in enumerate(f2.values, start=86):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=r_i, column=cols_3r[c_idx]).value = val

                        elif sheet_name == "4det1":
                            if t1_4d_dfs:
                                f1 = pd.concat(t1_4d_dfs, ignore_index=True)
                                for r_i, row in enumerate(f1.values, start=10):
                                    for idx, val in enumerate(row):
                                        if pd.notna(val): ts.cell(row=r_i, column=c4[idx]+1).value = val
                            if t2_4d_dfs:
                                f2 = pd.concat(t2_4d_dfs, ignore_index=True)
                                for r_i, row in enumerate(f2.values, start=553):
                                    for idx, val in enumerate(row):
                                        if pd.notna(val): ts.cell(row=r_i, column=c4[idx]+1).value = val

                        elif sheet_name == "4R":
                            # Бичих багануудын Excel дугаар: A,B,C (1,2,3), F,G,H,I (6,7,8,9), M,N (13,14)
                            cols_4r = [1, 2, 3, 6, 7, 8, 9, 13, 14]
                            
                            if r4_dfs:
                                f4r = pd.concat(r4_dfs, ignore_index=True)
                                # A6 мөрнөөс эхэлж бичих
                                for r_i, row in enumerate(f4r.values, start=6):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=r_i, column=cols_4r[c_idx]).value = val

                        elif sheet_name == "4R(i)":
                            # M багана нь 13 дахь багана (index 13)
                            start_col = 13 
                            
                            for c_idx, column_data in enumerate(r4_i_cols):
                                # Компани бүрээр багана ахиулах
                                current_col = start_col + c_idx
                                for r_idx, value in enumerate(column_data, start=4): # 4-р мөрнөөс эхэлнэ
                                    if pd.notna(value):
                                        ts.cell(row=r_idx, column=current_col).value = value

                        elif sheet_name == "5det1":
                            if det5_1_data:
                                ts = book[sheet_name]
                                
                                # ========================================
                                # 1. Sum утгуудыг тооцоолж бичих
                                # ========================================
                                
                                # E4:I5 хүртэлх нүднүүдийн sum (DataFrame индекс: мөр 1-2, багана 0-4)
                                for row_idx in range(1, 3):  # DataFrame-ийн 1, 2 мөр
                                    for col_idx in range(0, 5):  # Багана 0-4 (E-I)
                                        total = sum(
                                            float(df.iloc[row_idx, col_idx]) 
                                            for df in det5_1_data 
                                            if pd.notna(df.iloc[row_idx, col_idx]) and str(df.iloc[row_idx, col_idx]).replace('.','',1).replace('-','',1).isdigit()
                                        )
                                        # Excel мөр 4-5, багана E-I (5-9)
                                        ts.cell(row=row_idx + 3, column=col_idx + 5).value = total if total != 0 else None
                                
                                # F7:I15 хүртэлх нүднүүдийн sum (DataFrame индекс: мөр 4-12, багана 1-4)
                                for row_idx in range(4, 13):  # DataFrame-ийн 4-12 мөр
                                    for col_idx in range(1, 5):  # Багана 1-4 (F-I)
                                        total = sum(
                                            float(df.iloc[row_idx, col_idx]) 
                                            for df in det5_1_data 
                                            if pd.notna(df.iloc[row_idx, col_idx]) and str(df.iloc[row_idx, col_idx]).replace('.','',1).replace('-','',1).isdigit()
                                        )
                                        # Excel мөр 7-15, багана F-I (6-9)
                                        ts.cell(row=row_idx + 3, column=col_idx + 5).value = total if total != 0 else None
                                
                                # E22:E26 хүртэлх нүднүүдийн sum (DataFrame индекс: мөр 19-23, багана 0)
                                for row_idx in range(19, 24):  # DataFrame-ийн 19-23 мөр
                                    total = sum(
                                        float(df.iloc[row_idx, 0]) 
                                        for df in det5_1_data 
                                        if pd.notna(df.iloc[row_idx, 0]) and str(df.iloc[row_idx, 0]).replace('.','',1).replace('-','',1).isdigit()
                                    )
                                    # Excel мөр 22-26, багана E (5)
                                    ts.cell(row=row_idx + 3, column=5).value = total if total != 0 else None
                                
                                # ========================================
                                # 2. Компани бүрийн датаг N3-с эхлэн хуулах
                                # ========================================
                                
                                start_col = 14  # N багана = 14
                                
                                for comp_idx, df in enumerate(det5_1_data):
                                    # Компани бүр 5 багана эзлэх (E:I → 5 багана)
                                    base_col = start_col + (comp_idx * 5)
                                    
                                    # E3:I27 датаг хуулах (25 мөр × 5 багана)
                                    for row_idx in range(25):  # 0-24 (25 мөр)
                                        for col_idx in range(5):  # 0-4 (5 багана)
                                            value = df.iloc[row_idx, col_idx]
                                            if pd.notna(value):
                                                # Excel мөр: 3 + row_idx (3-р мөрнөөс эхэлнэ)
                                                # Excel багана: base_col + col_idx
                                                ts.cell(
                                                    row=3 + row_idx, 
                                                    column=base_col + col_idx
                                                ).value = value

                        elif sheet_name == "5det2":
                            if t1_5d_dfs:
                                f1 = pd.concat(t1_5d_dfs, ignore_index=True)
                                for r_i, row in enumerate(f1.values, start=8):
                                    for idx, val in enumerate(row):
                                        if pd.notna(val): ts.cell(row=r_i, column=c5[idx]+1).value = val
                            if t2_5d_dfs:
                                f2 = pd.concat(t2_5d_dfs, ignore_index=True)
                                for r_i, row in enumerate(f2.values, start=192):
                                    for idx, val in enumerate(row):
                                        if pd.notna(val): ts.cell(row=r_i, column=c5[idx]+1).value = val

                        elif sheet_name == "5R(i)":
                            for c_i, cv in enumerate(r5_i_cols):
                                for r_i, val in enumerate(cv, start=4):
                                    if pd.notna(val): ts.cell(row=r_i, column=16+c_i).value = val

                        elif sheet_name == "5R(ii)":
                            for c_i, cv in enumerate(r5_ii_cols):
                                for r_i, val in enumerate(cv, start=4):
                                    if pd.notna(val): ts.cell(row=r_i, column=12+c_i).value = val

                        elif sheet_name == "6det1":
                            if t1_6d_dfs:
                                f1 = pd.concat(t1_6d_dfs, ignore_index=True)
                                for r_i, row in enumerate(f1.values, start=10):
                                    for idx, val in enumerate(row):
                                        if pd.notna(val): ts.cell(row=r_i, column=c6[idx]+1).value = val
                            if t2_6d_dfs:
                                f2 = pd.concat(t2_6d_dfs, ignore_index=True)
                                for r_i, row in enumerate(f2.values, start=733):
                                    for idx, val in enumerate(row):
                                        if pd.notna(val): ts.cell(row=r_i, column=c6[idx]+1).value = val
                        
                        elif sheet_name == "6det2":
                            cols_6d2 = [1, 2, 3, 4, 7, 8, 9, 10, 11, 12]
                            
                            if t1_6d2_dfs:
                                f1 = pd.concat(t1_6d2_dfs, ignore_index=True)
                                for r_i, row in enumerate(f1.values, start=6):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=r_i, column=cols_6d2[c_idx]).value = val
                                            
                            if t2_6d2_dfs:
                                f2 = pd.concat(t2_6d2_dfs, ignore_index=True)
                                for r_i, row in enumerate(f2.values, start=150):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=r_i, column=cols_6d2[c_idx]).value = val

                        elif sheet_name == "6R":
                            # M багана нь 13 дахь багана (index 13)
                            start_col = 13 
                            
                            for c_idx, column_data in enumerate(r6_cols):
                                # Компани бүрээр багана ахиулах
                                current_col = start_col + c_idx
                                for r_idx, value in enumerate(column_data, start=4): # 4-р мөрнөөс эхэлнэ
                                    if pd.notna(value):
                                        ts.cell(row=r_idx, column=current_col).value = value

                        elif sheet_name == "7det1":
                            # Excel-ийн баганы дугаар (1-ээс эхэлсэн): 
                            # A=1, B=2, F=6, G=7, H=8, I=9, J=10, K=11, M=13, N=14, V=22, W=23, X=24, Y=25
                            # + AA=27, AB=28
                            # + AZ=52, BA=53, BB=54, BC=55, BD=56, BE=57, BF=58
                            cols_7d1 = [1, 2, 6, 7, 8, 9, 10, 11, 13, 14, 22, 23, 24, 25, 27, 28, 52, 53, 54, 55, 56, 57, 58]
                            
                            # Table 1: A10-аас эхэлж бичих
                            if t1_7d1_dfs:
                                f1 = pd.concat(t1_7d1_dfs, ignore_index=True)
                                # enumerate(..., start=10) нь мөрийн дугаарыг 10-аас эхлүүлнэ
                                for r_i, row in enumerate(f1.values, start=10):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            # cols_7d1[c_idx] нь тухайн утга Excel-ийн аль баганад очихыг заана
                                            ts.cell(row=r_i, column=cols_7d1[c_idx]).value = val
                                            
                            # Table 2: A28101-аас эхэлж бичих
                            if t2_7d1_dfs:
                                f2 = pd.concat(t2_7d1_dfs, ignore_index=True)
                                for r_i, row in enumerate(f2.values, start=28101):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=r_i, column=cols_7d1[c_idx]).value = val

                        elif sheet_name == "7det2":
                            # Table 1-ийг A6-аас эхлэн хуулах (start=6)
                            if t1_7d2_dfs:
                                f1 = pd.concat(t1_7d2_dfs, ignore_index=True)
                                # Баганын зураглал: A,B,C,D -> 1,2,3,4 болон H,I,J,K,L,M -> 8,9,10,11,12,13
                                col_map = [1, 2, 3, 4, 8, 9, 10, 11, 12, 13]
                                for r_i, row in enumerate(f1.values, start=6):
                                    for c_i, val in enumerate(row):
                                        target_col = col_map[c_i]
                                        if pd.notna(val): 
                                            ts.cell(row=r_i, column=target_col).value = val
                            
                            # Table 2-ийг A532-оос эхлэн хуулах (start=532)
                            if t2_7d2_dfs:
                                f2 = pd.concat(t2_7d2_dfs, ignore_index=True)
                                col_map = [1, 2, 3, 4, 8, 9, 10, 11, 12, 13]
                                for r_i, row in enumerate(f2.values, start=532):
                                    for c_i, val in enumerate(row):
                                        target_col = col_map[c_i]
                                        if pd.notna(val): 
                                            ts.cell(row=r_i, column=target_col).value = val

                        elif sheet_name == "7det3":
                            # Excel-ийн баганын дугаар: 
                            # A=1, B=2, C=3, D=4, E=5, F=6, G=7, H=8, J=10, K=11, M=13, N=14, P=16, Q=17, S=19, T=20, V=22, W=23, Y=25, Z=26
                            cols_7d3 = [1, 2, 3, 4, 5, 6, 7, 8, 10, 11, 13, 14, 16, 17, 19, 20, 22, 23, 25, 26]
                            
                            if t1_7d3_dfs:
                                f1 = pd.concat(t1_7d3_dfs, ignore_index=True)
                                # A9-өөс эхэлж бичих
                                for r_i, row in enumerate(f1.values, start=9):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=r_i, column=cols_7d3[c_idx]).value = val

                        elif sheet_name == "7R":
                            if r7_cols:
                                # Y баганаас эхэлнэ (A=1, ..., Y=25)
                                start_col = 25 
                                for c_idx, column_data in enumerate(r7_cols):
                                    current_col = start_col + c_idx
                                    # 3-р мөрнөөс эхлэн (F3-тай ижил түвшинд) доошоо бичих
                                    for r_idx, value in enumerate(column_data, start=3):
                                        if pd.notna(value):
                                            ts.cell(row=r_idx, column=current_col).value = value

                        elif sheet_name == "7R(i)":
                            # L багана нь 12 дахь багана (index 12)
                            start_col = 12 
                            
                            for c_idx, column_data in enumerate(r7_i_cols):
                                # Компани бүрээр багана ахиулах (L, M, N...)
                                current_col = start_col + c_idx
                                for r_idx, value in enumerate(column_data, start=2): # 2-р мөрнөөс эхэлнэ
                                    if pd.notna(value):
                                        ts.cell(row=r_idx, column=current_col).value = value

                        elif sheet_name == "7R(ii)":
                            # R багана нь 18 дахь багана (index 18)
                            start_col = 18 
                            
                            for comp_idx, data_dict in enumerate(r7_ii_data):
                                # Компани бүр 3 багана эзлэх тул үржих нь 3
                                base_col = start_col + (comp_idx * 3)
                                
                                # 1-р багана (E -> R, U, X...)
                                for r_idx, val in enumerate(data_dict['col_E'], start=5):
                                    if pd.notna(val): ts.cell(row=r_idx, column=base_col).value = val
                                    
                                # 2-р багана (I -> S, V, Y...)
                                for r_idx, val in enumerate(data_dict['col_I'], start=5):
                                    if pd.notna(val): ts.cell(row=r_idx, column=base_col + 1).value = val
                                    
                                # 3-р багана (M -> T, W, Z...)
                                for r_idx, val in enumerate(data_dict['col_M'], start=5):
                                    if pd.notna(val): ts.cell(row=r_idx, column=base_col + 2).value = val

                        elif sheet_name == "7R(iii)":
                            if r7iii_cols:
                                # T баганаас эхэлнэ (A=1, ..., T=20)
                                start_col = 20 
                                for c_idx, column_data in enumerate(r7iii_cols):
                                    current_col = start_col + c_idx
                                    # 4-р мөрнөөс эхлэн (D4-тэй ижил түвшинд) доошоо бичих
                                    for r_idx, value in enumerate(column_data, start=4):
                                        if pd.notna(value):
                                            ts.cell(row=r_idx, column=current_col).value = value

                        elif sheet_name == "8det1":
                            # Template-ийн баганууд: A=1, B=2, G=7, H=8, I=9, J=10, K=11, L=12
                            cols_8d1 = [1, 2, 7, 8, 9, 10, 11, 12]
                            
                            if t1_8d1_dfs:
                                f1 = pd.concat(t1_8d1_dfs, ignore_index=True)
                                start_row = 7
                                for r_i, row in enumerate(f1.values):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            # cols_8d1 жагсаалтын дагуу харгалзах баганад бичнэ
                                            ts.cell(row=start_row + r_i, column=cols_8d1[c_idx]).value = val

                        elif sheet_name == "8det2":
                            # Table 1: A7-оос эхэлж бичих
                            if t1_8d2_dfs:
                                f1 = pd.concat(t1_8d2_dfs, ignore_index=True)
                                for r_i, row in enumerate(f1.values, start=7):
                                    for c_i, val in enumerate(row, start=1):
                                        if pd.notna(val):
                                            ts.cell(row=r_i, column=c_i).value = val
                            
                            # Table 2: A78-аас эхэлж бичих
                            if t2_8d2_dfs:
                                f2 = pd.concat(t2_8d2_dfs, ignore_index=True)
                                for r_i, row in enumerate(f2.values, start=78):
                                    for c_i, val in enumerate(row, start=1):
                                        if pd.notna(val):
                                            ts.cell(row=r_i, column=c_i).value = val

                        elif sheet_name == "8R":
                            if r8_cols:
                                # O баганаас эхэлнэ (A=1, ..., O=15)
                                start_col = 15 
                                for c_idx, column_data in enumerate(r8_cols):
                                    current_col = start_col + c_idx
                                    # 2-р мөрнөөс эхлэн (E2-той ижил түвшинд) доошоо бичих
                                    for r_idx, value in enumerate(column_data, start=2):
                                        if pd.notna(value):
                                            ts.cell(row=r_idx, column=current_col).value = value

                        elif sheet_name == "8R(i)":
                            if r8i_cols:
                                # S баганаас эхэлнэ (A=1, ..., S=19)
                                start_col = 19 
                                for c_idx, column_data in enumerate(r8i_cols):
                                    current_col = start_col + c_idx
                                    # 5-р мөрнөөс эхлэн (D5-тай ижил түвшинд) доошоо бичих
                                    for r_idx, value in enumerate(column_data, start=5):
                                        if pd.notna(value):
                                            ts.cell(row=r_idx, column=current_col).value = value

                        elif sheet_name == "9det":
                            # 8 хүснэгтийн мэдээллийг жагсаалтаар тодорхойлох
                            tables_info = [
                                (t1_9d_dfs, 6),    # Table 1 эхлэх мөр: 6
                                (t2_9d_dfs, 72),   # Table 2 эхлэх мөр: 72
                                (t3_9d_dfs, 1088), # Table 3 эхлэх мөр: 1088
                                (t4_9d_dfs, 1262), # Table 4 эхлэх мөр: 1262
                                (t5_9d_dfs, 1296), # Table 5 эхлэх мөр: 1296
                                (t6_9d_dfs, 1355), # Table 6 эхлэх мөр: 1355
                                (t7_9d_dfs, 1368), # Table 7 эхлэх мөр: 1368
                                (t8_9d_dfs, 1377)  # Table 8 эхлэх мөр: 1377
                            ]
                            
                            for t_list, start_row in tables_info:
                                # ХАМГИЙН ЧУХАЛ: Жагсаалт хоосон биш, дотор нь DataFrame байгаа эсэхийг шалгах
                                # None утгуудыг шүүж, зөвхөн бодит датаг авна
                                valid_dfs = [df for df in t_list if df is not None and not df.empty]
                                
                                if valid_dfs:
                                    try:
                                        # Зөвхөн дата байгаа тохиолдолд нэгтгэнэ
                                        final_df = pd.concat(valid_dfs, ignore_index=True)
                                        
                                        # Template руу бичих
                                        for r_i, row in enumerate(final_df.values):
                                            for c_i, val in enumerate(row):
                                                if pd.notna(val):
                                                    # start_row-оос эхлэн утгуудыг нүдэнд оноох
                                                    ts.cell(row=start_row + r_i, column=c_i + 1).value = val
                                    except Exception as concat_err:
                                        # Тухайн нэг хүснэгт дээр алдаа гарвал бусдыг нь зогсоохгүй
                                        print(f"9det хүснэгт нэгтгэх алдаа (мөр {start_row}): {concat_err}")
                                        continue
                                else:
                                    # Хэрэв дата байхгүй бол тухайн хүснэгтийг зүгээр л алгасаад дараагийнх руу орно
                                    continue

                        elif sheet_name == "9R":
                            # N багана нь 14 дахь багана (index 14)
                            start_col = 14 
                            
                            for c_idx, column_data in enumerate(r9_cols):
                                # Компани бүрээр багана ахиулах (N, O, P...)
                                current_col = start_col + c_idx
                                for r_idx, value in enumerate(column_data, start=4): # 4-р мөрнөөс эхэлнэ
                                    if pd.notna(value):
                                        ts.cell(row=r_idx, column=current_col).value = value

                        elif sheet_name == "10det1":
                            tables_to_write = [
                                (t1_10d1_dfs, 13),
                                (t2_10d1_dfs, 253),
                                (t3_10d1_dfs, 371),
                                (t4_10d1_dfs, 420)
                            ]
                            
                            # Баганын хаягуудыг Excel-ийн A, B, F... хэлбэрээр тодорхойлох
                            # Энэ нь template-ийн яг заасан баганууд руу бичихэд хэрэгтэй
                            target_cols = [1, 2] + list(range(6, 13)) + list(range(15, 19)) + \
                                          list(range(21, 33)) + list(range(37, 52)) + list(range(54, 60))

                            for t_list, start_row in tables_to_write:
                                valid_dfs = [df for df in t_list if df is not None and not df.empty]
                                if valid_dfs:
                                    final_df = pd.concat(valid_dfs, ignore_index=True)
                                    for r_i, row in enumerate(final_df.values):
                                        for c_i, val in enumerate(row):
                                            if pd.notna(val):
                                                # target_cols-оос баганын дугаарыг авч бичнэ
                                                ts.cell(row=start_row + r_i, column=target_cols[c_i]).value = val

                        elif sheet_name == "10det2":
                            valid_dfs = [df for df in all_10d2_dfs if df is not None and not df.empty]
                            if valid_dfs:
                                final_df = pd.concat(valid_dfs, ignore_index=True)
                                start_row_tpl = 7
                                
                                # Бичих багануудын жагсаалт: 
                                # A (1), D (4), E (5), F (6), G (7), H (8), I (9), J (10), K (11), L (12), M (13)
                                # B (2), C (3) багануудыг алгасаж байна.
                                target_columns = [1] + list(range(4, 14)) 
                                
                                # DataFrame-ийн индексүүд:
                                # Index 0 = A (Company Code)
                                # Index 3 = D, Index 4 = E ... Index 12 = M
                                source_indices = [0] + list(range(3, 13))

                                for r_idx, row in enumerate(final_df.values):
                                    for i, src_idx in enumerate(source_indices):
                                        val = row[src_idx]
                                        if pd.notna(val):
                                            # target_columns[i] нь тухайн утга очих Excel-ийн багана
                                            ts.cell(row=start_row_tpl + r_idx, column=target_columns[i]).value = val

                        elif sheet_name == "10det3":
                            valid_dfs = [df for df in all_10d3_dfs if df is not None and not df.empty]
                            if valid_dfs:
                                final_df = pd.concat(valid_dfs, ignore_index=True)
                                
                                # Template-ийн A7 (Row 7, Column 1) -оос эхэлнэ
                                start_row_tpl = 7
                                
                                for r_idx, row in enumerate(final_df.values):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            # Column 1-ээс эхлэн (A, B, C... L) дараалан бичнэ
                                            ts.cell(row=start_row_tpl + r_idx, column=c_idx + 1).value = val

                        elif sheet_name == "Repo1":
                            # --- Table 1 Бичих ---
                            valid_t1 = [df for df in all_repo1_t1_dfs if not df.empty]
                            if valid_t1:
                                final_t1 = pd.concat(valid_t1, ignore_index=True)
                                # Target columns (Excel-ийн баганын дугаар 1-ээс эхэлнэ)
                                t1_target = [1, 2, 6, 7, 8, 9, 10, 13, 14, 19, 20, 21, 22, 23, 24, 25, 26, 28, 34, 35, 36, 37]
                                for r_idx, row in enumerate(final_t1.values):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=10 + r_idx, column=t1_target[c_idx]).value = val

                            # --- Table 2 Бичих ---
                            valid_t2 = [df for df in all_repo1_t2_dfs if not df.empty]
                            if valid_t2:
                                final_t2 = pd.concat(valid_t2, ignore_index=True)
                                # Target columns: A:B, F:J, M:N, AB, AH:AI
                                t2_target = [1, 2, 6, 7, 8, 9, 10, 13, 14, 28, 34, 35]
                                for r_idx, row in enumerate(final_t2.values):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=98 + r_idx, column=t2_target[c_idx]).value = val

                        elif sheet_name == "Repo2":
                            # Бичих багануудын Excel дугаар (Target Column indices)
                            # A:B(1,2), F:J(6,7,8,9,10), M:P(13,14,15,16), S:Z(19...26), AE:AJ(31...36), AL:AQ(38...43)
                            repo2_target = [1, 2] + list(range(6, 11)) + list(range(13, 17)) + \
                                          list(range(19, 27)) + list(range(31, 37)) + list(range(38, 44))

                            # --- Table 1 Бичих (A13-аас) ---
                            valid_t1 = [df for df in all_repo2_t1_dfs if not df.empty]
                            if valid_t1:
                                final_t1 = pd.concat(valid_t1, ignore_index=True)
                                for r_idx, row in enumerate(final_t1.values):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=13 + r_idx, column=repo2_target[c_idx]).value = val

                            # --- Table 2 Бичих (A36-аас) ---
                            valid_t2 = [df for df in all_repo2_t2_dfs if not df.empty]
                            if valid_t2:
                                final_t2 = pd.concat(valid_t2, ignore_index=True)
                                for r_idx, row in enumerate(final_t2.values):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=36 + r_idx, column=repo2_target[c_idx]).value = val

                        elif sheet_name == "Repo":
                            if all_repo_cols:
                                # K4 нь Column 11, Row 4 юм
                                start_col_tpl = 11
                                start_row_tpl = 4
                                
                                # Компани бүрийн өгөгдлийг багана баганаар бичих
                                for c_idx, col_data in enumerate(all_repo_cols):
                                    # c_idx=0 бол K багана, c_idx=1 бол L багана...
                                    target_col = start_col_tpl + c_idx
                                    
                                    for r_idx, val in enumerate(col_data):
                                        if pd.notna(val):
                                            ts.cell(row=start_row_tpl + r_idx, column=target_col).value = val

                        elif sheet_name == "10R(i)":
                            # --- Table 1 Бичих (A7-оос) ---
                            valid_t1 = [df for df in all_10ri_t1_dfs if not df.empty]
                            if valid_t1:
                                final_t1 = pd.concat(valid_t1, ignore_index=True)
                                for r_idx, row in enumerate(final_t1.values):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            ts.cell(row=7 + r_idx, column=c_idx + 1).value = val

                            # --- Table 2 Бичих (A76-аас) ---
                            valid_t2 = [df for df in all_10ri_t2_dfs if not df.empty]
                            if valid_t2:
                                final_t2 = pd.concat(valid_t2, ignore_index=True)
                                for r_idx, row in enumerate(final_t2.values):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            # A76 нь Row 76, Column 1
                                            ts.cell(row=76 + r_idx, column=c_idx + 1).value = val

                        elif sheet_name == "10R(ii)":
                            valid_dfs = [df for df in all_10rii_dfs if df is not None and not df.empty]
                            if valid_dfs:
                                final_df = pd.concat(valid_dfs, ignore_index=True)
                                
                                # Template-ийн A10 (Row 10, Column 1) -аас эхэлнэ
                                start_row_tpl = 10
                                
                                for r_idx, row in enumerate(final_df.values):
                                    for c_idx, val in enumerate(row):
                                        if pd.notna(val):
                                            # Column 1-ээс эхлэн (A, B... J) дараалан бичнэ
                                            ts.cell(row=start_row_tpl + r_idx, column=c_idx + 1).value = val

                        elif sheet_name == "10R(iii)":
                            if all_10riii_cols:
                                # W4 нь Column 23, Row 4 юм
                                start_col_tpl = 23
                                start_row_tpl = 4
                                
                                # Компани бүрийн өгөгдлийг багана баганаар бичих
                                for c_idx, col_data in enumerate(all_10riii_cols):
                                    # c_idx=0 бол W багана, c_idx=1 бол X багана...
                                    target_col = start_col_tpl + c_idx
                                    
                                    for r_idx, val in enumerate(col_data):
                                        if pd.notna(val):
                                            ts.cell(row=start_row_tpl + r_idx, column=target_col).value = val

                        elif sheet_name == "10R(iv)":
                            # Бичих тохиргоонууд: (Датаны жагсаалт, Template-ийн эхлэх Row)
                            write_configs = [
                                (all_10riv_t1_dfs, 11),
                                (all_10riv_t2_dfs, 98),
                                (all_10riv_t3_dfs, 179)
                            ]

                            for dfs_list, start_row_tpl in write_configs:
                                valid_dfs = [df for df in dfs_list if not df.empty]
                                if valid_dfs:
                                    final_df = pd.concat(valid_dfs, ignore_index=True)
                                    for r_idx, row in enumerate(final_df.values):
                                        for c_idx, val in enumerate(row):
                                            if pd.notna(val):
                                                # Column 1-ээс эхлэн (A, B, C, D, E) дараалан бичнэ
                                                ts.cell(row=start_row_tpl + r_idx, column=c_idx + 1).value = val

                        elif sheet_name == "10R(v)":
                            if all_10rv_cols:
                                start_col_v = 13  # M багана
                                
                                # Нийлбэрийг хадгалах хоосон жагсаалт (15 нүд: Row 7-оос 21 хүртэл)
                                # 10R(v) F7:F21 нь индексээр 3-аас (F7) 17-р (F21) элементүүд байна.
                                totals_f7_f21 = [0.0] * 15 
                                
                                for comp_idx, col_data in enumerate(all_10rv_cols):
                                    current_col_v = start_col_v + comp_idx
                                    
                                    for row_idx, val in enumerate(col_data):
                                        # A. Хуучин логик: M4-өөс эхлэн хажуу тийш цувуулах
                                        if pd.notna(val):
                                            ts.cell(row=4 + row_idx, column=current_col_v).value = val
                                        
                                        # B. Шинэ логик: F7:F21-ийн нийлбэрийг тооцоолох
                                        # col_data-ийн index 3-аас эхэлж байгаа нь F7 мөр юм.
                                        if 3 <= row_idx <= 17:
                                            try:
                                                num_val = float(val) if pd.notna(val) else 0.0
                                                totals_f7_f21[row_idx - 3] += num_val
                                            except (ValueError, TypeError):
                                                pass

                                # C. Template-ийн F7:F21 нүднүүдэд нийлбэр дүнг бичих
                                # F багана нь 6 дахь багана (column=6)
                                for i, total_val in enumerate(totals_f7_f21):
                                    ts.cell(row=7 + i, column=6).value = total_val

                        elif sheet_name == "11det1":
                            if all_11det1_dfs:
                                final_df_11d1 = pd.concat(all_11det1_dfs, ignore_index=True)
                                
                                # Template-ийн багануудын нэршил (A:AM хүртэлх харгалзах баганууд)
                                # Бид зөвхөн өөрийн сонгосон баганууд руугаа бичнэ
                                target_col_indices = list(range(1, 19)) + [20, 21, 22, 23, 24] + [26, 27] + \
                                                     [29, 30] + [32, 33] + [35, 36] + [38, 39]
                                
                                start_row = 9
                                for r_idx, row_data in enumerate(final_df_11d1.values):
                                    for c_idx, value in enumerate(row_data):
                                        if pd.notna(value):
                                            # Excel-ийн баганын индекс (1-ээс эхэлнэ)
                                            actual_col = target_col_indices[c_idx]
                                            ts.cell(row=start_row + r_idx, column=actual_col).value = value

                        elif sheet_name == "11det2":
                            if all_11det2_dfs:
                                final_df_11d2 = pd.concat(all_11det2_dfs, ignore_index=True)
                                
                                # Template-ийн харгалзах багануудын индекс (Excel багана 1-ээс эхэлнэ)
                                target_col_indices = [1, 2] + list(range(6, 11)) + [13, 14] + \
                                                     list(range(17, 29)) + list(range(34, 40)) + \
                                                     list(range(41, 47))
                                
                                start_row = 12
                                for r_idx, row_data in enumerate(final_df_11d2.values):
                                    for c_idx, value in enumerate(row_data):
                                        if pd.notna(value):
                                            # Excel-ийн баганыг зааж өгч бичнэ
                                            actual_col = target_col_indices[c_idx]
                                            ts.cell(row=start_row + r_idx, column=actual_col).value = value

                        elif sheet_name == "11R":
                            if all_11r_cols:
                                # S баганаас эхэлнэ (S=19, T=20, U=21)
                                start_col_11r = 19 
                                
                                for comp_idx, triple_cols in enumerate(all_11r_cols):
                                    # Компани бүрт 3 багана оноогдоно (S,T,U дараа нь V,W,X гэх мэт)
                                    base_col = start_col_11r + (comp_idx * 3)
                                    
                                    for sub_idx, col_data in enumerate(triple_cols):
                                        current_col = base_col + sub_idx
                                        for row_idx, val in enumerate(col_data):
                                            # 3-р мөрнөөс (row=3) эхлэн доошоо бичнэ
                                            if pd.notna(val):
                                                ts.cell(row=3 + row_idx, column=current_col).value = val

                        elif sheet_name == "11R(i)":
                            if all_11ri_cols:
                                # O баганаас эхэлнэ (O=15, P=16)
                                start_col_11ri = 15 
                                
                                for comp_idx, pair_cols in enumerate(all_11ri_cols):
                                    # Компани бүрт 2 багана оноогдоно (O,P дараа нь Q,R гэх мэт)
                                    # Алхам нь 2-оор үсэрнэ
                                    base_col = start_col_11ri + (comp_idx * 2)
                                    
                                    for sub_idx, col_data in enumerate(pair_cols):
                                        current_col = base_col + sub_idx
                                        for row_idx, val in enumerate(col_data):
                                            # 4-р мөрнөөс (row=4) эхлэн доошоо бичнэ
                                            if pd.notna(val):
                                                ts.cell(row=4 + row_idx, column=current_col).value = val

                        elif sheet_name == "12det":
                            if all_12det_dfs:
                                final_df_12d = pd.concat(all_12det_dfs, ignore_index=True)
                                
                                # Template-ийн харгалзах багануудын индекс (Excel-ийн багана 1-ээс эхэлнэ)
                                # AH=34, AI=35, AJ=36, AK=37, AL=38, AM=39, AN=40
                                target_col_indices = [1, 2] + list(range(7, 11)) + [12] + \
                                                     list(range(18, 24)) + [26, 27] + \
                                                     list(range(34, 41))
                                
                                start_row = 12
                                for r_idx, row_data in enumerate(final_df_12d.values):
                                    for c_idx, value in enumerate(row_data):
                                        if pd.notna(value):
                                            # target_col_indices-ээс тухайн утга Excel-ийн аль баганад очихыг авна
                                            try:
                                                actual_col = target_col_indices[c_idx]
                                                ts.cell(row=start_row + r_idx, column=actual_col).value = value
                                            except IndexError:
                                                # Хэрэв дата болон баганын индексийн тоо зөрвөл алгасах
                                                continue

                        elif sheet_name == "12R":
                            if all_12r_cols:
                                # M баганаас эхэлнэ (M=13, N=14, O=15...)
                                start_col_12r = 13 
                                
                                for comp_idx, col_data in enumerate(all_12r_cols):
                                    current_col_12r = start_col_12r + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 2-р мөрнөөс (row=2) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=2 + row_idx, column=current_col_12r).value = val

                        elif sheet_name == "13det":
                            # Target багануудын индекс (A, B, C, E, F, G, H, I, J, K, L)
                            target_indices = [1, 2, 3, 5, 6, 7, 8, 9, 10, 11, 12]
                            
                            # 1. Table 1 нэгтгэж A7-оос эхлэн бичих
                            if all_13det_t1_dfs:
                                final_t1 = pd.concat(all_13det_t1_dfs, ignore_index=True)
                                start_row_t1 = 7
                                for r_idx, row_data in enumerate(final_t1.values):
                                    for c_idx, val in enumerate(row_data):
                                        if pd.notna(val):
                                            ts.cell(row=start_row_t1 + r_idx, column=target_indices[c_idx]).value = val
                            
                            # 2. Table 2 нэгтгэж A131-ээс эхлэн бичих
                            if all_13det_t2_dfs:
                                final_t2 = pd.concat(all_13det_t2_dfs, ignore_index=True)
                                start_row_t2 = 131
                                for r_idx, row_data in enumerate(final_t2.values):
                                    for c_idx, val in enumerate(row_data):
                                        if pd.notna(val):
                                            ts.cell(row=start_row_t2 + r_idx, column=target_indices[c_idx]).value = val
                        
                        elif sheet_name == "13R":
                            if all_13r_cols:
                                # J баганаас эхэлнэ (J=10, K=11, L=12...)
                                start_col_13r = 10 
                                
                                for comp_idx, col_data in enumerate(all_13r_cols):
                                    current_col_13r = start_col_13r + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 4-р мөрнөөс (row=4) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=4 + row_idx, column=current_col_13r).value = val

                        elif sheet_name == "13R(i)":
                            if all_13ri_cols:
                                # Y баганаас эхэлнэ (Y=25, Z=26, AA=27...)
                                start_col_13ri = 25 
                                
                                for comp_idx, col_data in enumerate(all_13ri_cols):
                                    current_col_13ri = start_col_13ri + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 4-р мөрнөөс (row=4) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=4 + row_idx, column=current_col_13ri).value = val

                        elif sheet_name == "14R":
                            if all_14r_cols:
                                # J баганаас эхэлнэ (J=10, K=11, L=12...)
                                start_col_14r = 10 
                                
                                for comp_idx, col_data in enumerate(all_14r_cols):
                                    current_col_14r = start_col_14r + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 3-р мөрнөөс (row=3) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=3 + row_idx, column=current_col_14r).value = val

                        elif sheet_name.strip() == "14R(i)":
                            write_configs = [
                                (all_14r1_t1, 6), (all_14r1_t2, 63),
                                (all_14r1_t3, 118), (all_14r1_t4, 129)
                            ]
                            
                            for table_list, start_row in write_configs:
                                current_write_row = start_row
                                if table_list:
                                    # Компани бүрийн датаг цувуулж бичих
                                    for df_part, comp_code in table_list:
                                        for r_idx, row_values in enumerate(df_part.values):
                                            # A:L баганыг бичих (A багана оригиналь утгаараа байна)
                                            for c_idx, value in enumerate(row_values):
                                                if pd.notna(value):
                                                    ts.cell(row=current_write_row, column=c_idx + 1).value = value
                                            
                                            # W багана (23) дээр компанийн кодыг бичих
                                            ts.cell(row=current_write_row, column=23).value = comp_code
                                            
                                            current_write_row += 1

                        elif sheet_name == "17det1":
                            if all_17det1_dfs:
                                final_df_17d1 = pd.concat(all_17det1_dfs, ignore_index=True)
                                
                                # Template-ийн харгалзах багануудын индекс (Excel-ийн багана 1-ээс эхэлнэ)
                                target_col_indices = [1, 2, 3] + list(range(5, 13))
                                
                                start_row = 4
                                for r_idx, row_data in enumerate(final_df_17d1.values):
                                    for c_idx, value in enumerate(row_data):
                                        if pd.notna(value):
                                            # Заасан баганууд руу датаг бичнэ
                                            actual_col = target_col_indices[c_idx]
                                            ts.cell(row=start_row + r_idx, column=actual_col).value = value

                        elif sheet_name == "17det2":
                            if all_17det2_dfs:
                                final_df_17d2 = pd.concat(all_17det2_dfs, ignore_index=True)
                                
                                # Template-ийн харгалзах багануудын индекс (Excel-ийн багана 1-ээс эхэлнэ)
                                # A:F (1:7), H:L (8:13), O:T (15:21)
                                target_col_indices = list(range(1, 7)) + list(range(8, 13)) + list(range(15, 21))
                                
                                start_row = 7
                                for r_idx, row_data in enumerate(final_df_17d2.values):
                                    for c_idx, value in enumerate(row_data):
                                        if pd.notna(value):
                                            # Заасан баганууд руу датаг бичнэ
                                            actual_col = target_col_indices[c_idx]
                                            ts.cell(row=start_row + r_idx, column=actual_col).value = value

                        elif sheet_name == "17R":
                            if all_17r_cols:
                                # L баганаас эхэлнэ (L=12, M=13, N=14...)
                                start_col_17r = 12 
                                
                                for comp_idx, col_data in enumerate(all_17r_cols):
                                    current_col_17r = start_col_17r + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 3-р мөрнөөс (row=3) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=3 + row_idx, column=current_col_17r).value = val

                        elif sheet_name.strip() == "17R(ii)":
                            if all_17rii_data:
                                # N баганаас эхэлнэ (N=14, Q=17, T=20...)
                                start_col_17rii = 14 
                                
                                for comp_idx, comp_pack in enumerate(all_17rii_data):
                                    # Компани бүр 3 багана эзлэх тул: comp_idx * 3
                                    base_col = start_col_17rii + (comp_idx * 3)
                                    
                                    # 1. Компанийн кодыг 1-р мөрөнд бичих (N1, Q1, T1...)
                                    ts.cell(row=1, column=base_col).value = comp_pack['code']
                                    
                                    # 2. 3 багана датаг хуулах (2-р мөрнөөс эхлэн)
                                    for sub_col_idx, col_data in enumerate(comp_pack['columns']):
                                        target_col = base_col + sub_col_idx
                                        for row_idx, val in enumerate(col_data):
                                            if pd.notna(val):
                                                ts.cell(row=2 + row_idx, column=target_col).value = val
                                            else:
                                                ts.cell(row=2 + row_idx, column=target_col).value = None

                        elif sheet_name == "ICT Report Raw Summary":
                            if all_ict_summary_dfs:
                                # Бүх датаг нэгтгэх
                                final_ict_df = pd.concat(all_ict_summary_dfs, ignore_index=True)
                                
                                start_row_ict = 6 # Template-ийн F6 мөр
                                start_col_ict = 6 # Template-ийн F багана (6 дахь багана)
                                
                                for r_idx, row_data in enumerate(final_ict_df.values):
                                    for c_idx, val in enumerate(row_data):
                                        # Бүх утгыг бичнэ (Хоосон байсан ч хамаагүй)
                                        actual_row = start_row_ict + r_idx
                                        actual_col = start_col_ict + c_idx
                                        ts.cell(row=actual_row, column=actual_col).value = val

                        elif sheet_name.strip() == "18det":
                            if all_18det_cols:
                                # O баганаас эхэлнэ (O=15, P=16, Q=17...)
                                start_col_18det = 15 
                                
                                for comp_idx, col_data in enumerate(all_18det_cols):
                                    current_col_18det = start_col_18det + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 1-р мөрнөөс (row=1) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=1 + row_idx, column=current_col_18det).value = val

                        elif sheet_name.strip() == "19R":
                            if all_19r_cols:
                                # O баганаас эхэлнэ (O=15, P=16, Q=17...)
                                start_col_19r = 15 
                                
                                for comp_idx, col_data in enumerate(all_19r_cols):
                                    current_col_19r = start_col_19r + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 4-р мөрнөөс (row=4) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=4 + row_idx, column=current_col_19r).value = val

                        elif sheet_name.strip() == "20R":
                            if all_20r_cols:
                                # O баганаас эхэлнэ (O=15, P=16, Q=17...)
                                start_col_20r = 15 
                                
                                for comp_idx, col_data in enumerate(all_20r_cols):
                                    current_col_20r = start_col_20r + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 3-р мөрнөөс (row=3) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=3 + row_idx, column=current_col_20r).value = val

                        elif sheet_name.strip() == "20R(i)":
                            if all_20ri_cols:
                                # J баганаас эхэлнэ (J=10, K=11, L=12...)
                                start_col_20ri = 10 
                                
                                for comp_idx, col_data in enumerate(all_20ri_cols):
                                    current_col_20ri = start_col_20ri + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 3-р мөрнөөс (row=3) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=3 + row_idx, column=current_col_20ri).value = val

                        elif sheet_name.strip() == "21R":
                            if all_21r_cols:
                                # M баганаас эхэлнэ (M=13, N=14, O=15...)
                                start_col_21r = 13 
                                
                                for comp_idx, col_data in enumerate(all_21r_cols):
                                    current_col_21r = start_col_21r + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 3-р мөрнөөс (row=3) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=3 + row_idx, column=current_col_21r).value = val

                        elif sheet_name.strip() == "21R(i)":
                            if all_21ri_cols:
                                # H баганаас эхэлнэ (H=8, I=9, J=10...)
                                start_col_21ri = 8 
                                
                                for comp_idx, col_data in enumerate(all_21ri_cols):
                                    current_col_21ri = start_col_21ri + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # Template-ийн 3-р мөрнөөс (row=3) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=3 + row_idx, column=current_col_21ri).value = val

                        elif sheet_name.strip() == "22R(i)":
                            if all_22ri_cols:
                                # M баганаас эхэлнэ (M=13, N=14, O=15...)
                                start_col_22ri = 13 
                                
                                for comp_idx, col_data in enumerate(all_22ri_cols):
                                    current_col_22ri = start_col_22ri + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # Template-ийн 5-р мөрнөөс (row=5) эхлэн доошоо бичнэ
                                        if pd.notna(val):
                                            ts.cell(row=5 + row_idx, column=current_col_22ri).value = val

                        elif sheet_name.strip() == "22R(vi)":
                            if all_22rvi_cols:
                                # P баганаас эхэлнэ (P=16, Q=17, R=18, S=19...)
                                start_col_22rvi = 16 
                                
                                for comp_idx, (f_data, h_data) in enumerate(all_22rvi_cols):
                                    # Компани бүрт 2 багана: эхнийх нь P+0, P+2, P+4...
                                    # Дараагийнх нь Q+0, Q+2, Q+4...
                                    col_f_target = start_col_22rvi + (comp_idx * 2)
                                    col_h_target = col_f_target + 1
                                    
                                    for row_idx in range(len(f_data)):
                                        # 6-р мөрнөөс (row=6) эхлэн доошоо бичнэ
                                        # F датаг эхний баганад
                                        if pd.notna(f_data[row_idx]):
                                            ts.cell(row=6 + row_idx, column=col_f_target).value = f_data[row_idx]
                                        # H датаг дараагийн баганад
                                        if pd.notna(h_data[row_idx]):
                                            ts.cell(row=6 + row_idx, column=col_h_target).value = h_data[row_idx]
                            
                        elif sheet_name.strip() == "24det":
                            if all_24det_dfs:
                                final_df_24d = pd.concat(all_24det_dfs, ignore_index=True)
                                
                                start_row = 7
                                # A:L баганууд нь 1-ээс 12 дахь багана байна
                                for r_idx, row_data in enumerate(final_df_24d.values):
                                    for c_idx, value in enumerate(row_data):
                                        if pd.notna(value):
                                            ts.cell(row=start_row + r_idx, column=c_idx + 1).value = value

                        elif sheet_name.strip() == "24R(i)":
                            if all_24ri_sum_data:
                                # 1. Зөвхөн заасан Range-үүдэд SUM бодож бичих
                                # zip-ээр багана бүрийн тоон утгыг нэмнэ
                                sum_totals = [sum(values) for values in zip(*all_24ri_sum_data)]
                                
                                # Excel Row 6-11 (Index 2-7) болон Row 18-20 (Index 14-16)
                                target_indices = list(range(2, 8)) + list(range(14, 17))
                                
                                for i in target_indices:
                                    total_val = sum_totals[i]
                                    # Хэрэв нийлбэр нь 0 бол хоосон үлдээх эсвэл 0-ээр нь бичиж болно
                                    ts.cell(row=4 + i, column=6).value = total_val
                                
                                # 2. M баганаас эхлэн компани бүрээр цувуулах (Энд оригиналь утгаар нь)
                                start_col_24ri = 13 # M = 13
                                for comp_idx, col_data in enumerate(all_24ri_cols):
                                    current_col = start_col_24ri + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # pd.isna(val) бол нүдийг хоосон үлдээнэ, үгүй бол утгыг нь бичнэ
                                        target_cell = ts.cell(row=4 + row_idx, column=current_col)
                                        if pd.notna(val):
                                            target_cell.value = val
                                        else:
                                            target_cell.value = None # Хоосон байвал None (Excel дээр Empty харагдана)

                        elif sheet_name.strip() == "24R(ii)":
                            if all_24rii_cols:
                                # V баганаас эхэлнэ (V=22, W=23, X=24...)
                                start_col_24rii = 22 
                                
                                for comp_idx, col_data in enumerate(all_24rii_cols):
                                    current_col = start_col_24rii + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 5-р мөрнөөс (row=5) эхлэн доошоо бичнэ
                                        target_cell = ts.cell(row=5 + row_idx, column=current_col)
                                        if pd.notna(val):
                                            target_cell.value = val
                                        else:
                                            target_cell.value = None # Хоосон бол None үлдээнэ

                        elif sheet_name.strip() == "25det1":
                            if all_25det1_dfs:
                                final_df_25d1 = pd.concat(all_25det1_dfs, ignore_index=True)
                                
                                start_row = 7
                                # Бичих багануудын Excel дэх дугаар (A=1, B=2, D=4, E=5, G=7, J=10, K=11, L=12)
                                target_cols = [1, 2, 4, 5, 7, 10, 11, 12]
                                
                                for r_idx, row_values in enumerate(final_df_25d1.values):
                                    for c_idx, value in enumerate(row_values):
                                        if pd.notna(value):
                                            # target_cols-оос харгалзах баганын дугаарыг авна
                                            ts.cell(row=start_row + r_idx, column=target_cols[c_idx]).value = value

                        elif sheet_name.strip() == "25det2":
                            if all_25det2_dfs:
                                final_df_25d2 = pd.concat(all_25det2_dfs, ignore_index=True)
                                
                                start_row = 7
                                # Бичих багануудын Excel индекс (A=1, B=2, D=4, F=6, G=7)
                                target_cols = [1, 2, 4, 6, 7]
                                
                                for r_idx, row_values in enumerate(final_df_25d2.values):
                                    for c_idx, value in enumerate(row_values):
                                        if pd.notna(value):
                                            ts.cell(row=start_row + r_idx, column=target_cols[c_idx]).value = value

                        elif sheet_name.strip() == "25R(i)":
                            if all_25ri_cols:
                                # AA баганаас эхэлнэ (AA=27, AB=28, AC=29...)
                                start_col_25ri = 27 
                                
                                for comp_idx, col_data in enumerate(all_25ri_cols):
                                    current_col = start_col_25ri + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # 5-р мөрнөөс (row=5) эхлэн доошоо бичнэ
                                        target_cell = ts.cell(row=5 + row_idx, column=current_col)
                                        if pd.notna(val):
                                            target_cell.value = val
                                        else:
                                            target_cell.value = None # Хоосон бол None үлдээнэ

                        elif sheet_name.strip() == "25R(ii)":
                            if all_25rii_sum_data:
                                # 1. F7:F14 хүрээнд нийлбэр (SUM) бодож бичих
                                # zip-ээр багана бүрийн тоон утгыг нэмнэ
                                sum_totals = [sum(values) for values in zip(*all_25rii_sum_data)]
                                
                                # F4-өөс эхлэн уншсан тул:
                                # F7 = индекс 3, F14 = индекс 10. (Range 3-аас 11 хүртэл)
                                for i in range(3, 11):
                                    total_val = sum_totals[i]
                                    ts.cell(row=4 + i, column=6).value = total_val # Column 6 = F
                                
                                # 2. N4 баганаас эхлэн компани бүрээр цувуулах
                                start_col_25rii = 14 # N = 14
                                for comp_idx, col_data in enumerate(all_25rii_cols):
                                    current_col = start_col_25rii + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        target_cell = ts.cell(row=4 + row_idx, column=current_col)
                                        if pd.notna(val):
                                            target_cell.value = val
                                        else:
                                            target_cell.value = None

                        elif sheet_name.strip() == "26":
                            if all_26_sum_data:
                                # 1. C3:C9 хүрээнд нийлбэр (SUM) бодож бичих
                                # zip-ээр багана бүрийн тоон утгыг нэмнэ
                                sum_totals = [sum(values) for values in zip(*all_26_sum_data)]
                                
                                # C2-оос эхлэн уншсан тул:
                                # C3 = индекс 1, C9 = индекс 7. (Range 1-ээс 8 хүртэл)
                                for i in range(1, 8):
                                    total_val = sum_totals[i]
                                    ts.cell(row=2 + i, column=3).value = total_val # Column 3 = C
                                
                                # 2. G2 баганаас эхлэн компани бүрээр цувуулах
                                start_col_26 = 7 # G = 7
                                for comp_idx, col_data in enumerate(all_26_cols):
                                    current_col = start_col_26 + comp_idx
                                    for row_idx, val in enumerate(col_data):
                                        # Бүх утгыг (хоосон байсан ч 0 болгосон) бичнэ
                                        ts.cell(row=2 + row_idx, column=current_col).value = val

                final_io = io.BytesIO()
                book.save(final_io)
                final_io.seek(0)
                st.session_state.final_output = final_io
                
                status_text.empty()
                total_time = time.time() - start_time
                timer_text.markdown(f"✅ **Амжилттай дууслаа!** Нийт хугацаа: {total_time:.1f} сек")
                st.success("Нэгтгэх процесс дууслаа.")

                if error_logs:
                    with st.expander("⚠️ Боловсруулалтын явцад гарсан алдаанууд"):
                        for log in error_logs:
                            st.warning(log)

            except Exception as e:
                # Алдааг дэлгэц дээр хэвлэх, гэхдээ rerun хийхгүй
                st.error(f"⚠️ Нэгтгэх явцад алдаа гарлаа: {e}")
                # Туршлагатай хөгжүүлэгчид traceback-ийг консол дээр харахыг хүсвэл:
                print(traceback.format_exc())

            finally:
                st.session_state.is_processing = False
                import gc
                if 'template_bytes' in locals():
                    del template_bytes
                if 'book' in locals():
                    del book
                gc.collect()

# 3. "Татах" процесс
if st.session_state.final_output is not None:
    st.divider()
    dcol1, dcol2 = st.columns([1, 5])
    with dcol1:
        st.download_button(
            label="Татах",
            data=st.session_state.final_output,
            file_name="FRP_Consolidated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
    with dcol2:
        st.info("Нэгтгэсэн файлаа эндээс татаж авна уу.")